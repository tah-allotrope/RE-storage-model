"""
Extract reference KPIs from Excel financial model files.

Reads pre-calculated values from known cell positions in the Excel model
using openpyxl with data_only=True (cached formula results). Writes a
JSON reference file per project for use by regression tests.

Usage:
    python scripts/extract_excel_kpis.py tests/data/projects/project_01.xlsx
    python scripts/extract_excel_kpis.py tests/data/projects/*.xlsx

Output:
    tests/data/references/<stem>.json  (one per input file)

IMPORTANT: openpyxl data_only=True reads cached values. If the Excel file
was saved without recalculating, some cells may be None. Re-open the file
in Excel, press Ctrl+Shift+F9, save, then re-run this script.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Cell map: (sheet_name, cell_address) -> kpi_key
# These positions are derived from model_architecture.md §D.4 and the
# analysis_report.md Financial sheet section.
# ---------------------------------------------------------------------------

FINANCIAL_KPIS: dict[str, tuple[str, str]] = {
    "project_irr": ("Financial", "G123"),
    "equity_irr": ("Financial", "G136"),
    "unlevered_irr": ("Financial", "G189"),
    "npv_usd": ("Financial", "G193"),
}

# Intermediate KPIs for layer-by-layer debugging.
# These use row-label search because row positions may shift across files.
INTERMEDIATE_LABEL_KPIS: dict[str, dict[str, str]] = {
    # sheet -> { kpi_key: row_label_substring }
    # We search column B/C for the label and read the value from a known column.
}

# ---------------------------------------------------------------------------
# Calc sheet column statistics (read from the 8760-row hourly simulation)
# ---------------------------------------------------------------------------

CALC_COLUMN_STATS: dict[str, tuple[str, str]] = {
    # kpi_key: (column_letter, aggregation)
    # Column F = SolarGen_kW, Column M = SoC_kWh
    "calc_solar_gen_sum_kwh": ("F", "sum"),
    "calc_soc_min_kwh": ("M", "min"),
    "calc_soc_max_kwh": ("M", "max"),
}

HEADER_ROW = 1
DATA_START_ROW = 2


def _safe_float(value: Any) -> float | None:
    """Convert a cell value to float, returning None if not possible."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def extract_financial_kpis(wb: Workbook) -> dict[str, float | None]:
    """
    Extract final financial KPIs from known cell positions.

    Why: These cells contain XIRR/XNPV results that are the ultimate
    validation targets for the Python model.
    """
    results: dict[str, float | None] = {}
    for kpi_key, (sheet_name, cell_address) in FINANCIAL_KPIS.items():
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found — skipping %s", sheet_name, kpi_key)
            results[kpi_key] = None
            continue
        ws = wb[sheet_name]
        value = ws[cell_address].value
        results[kpi_key] = _safe_float(value)
        if results[kpi_key] is None:
            logger.warning(
                "KPI '%s' at %s!%s is None — Excel may need recalculation",
                kpi_key,
                sheet_name,
                cell_address,
            )
    return results


def extract_calc_column_stats(wb: Workbook) -> dict[str, float | None]:
    """
    Extract summary statistics from the Calc sheet's hourly columns.

    Why: These intermediate physics results let us isolate mismatches
    to the physics layer before checking settlement/financial layers.
    """
    results: dict[str, float | None] = {}
    sheet_name = "Calc"
    if sheet_name not in wb.sheetnames:
        logger.warning("Sheet '%s' not found — skipping Calc stats", sheet_name)
        return {k: None for k in CALC_COLUMN_STATS}

    ws = wb[sheet_name]
    max_row = ws.max_row or 1

    for kpi_key, (col_letter, agg) in CALC_COLUMN_STATS.items():
        values: list[float] = []
        for row in range(DATA_START_ROW, max_row + 1):
            cell_val = ws[f"{col_letter}{row}"].value
            num = _safe_float(cell_val)
            if num is not None:
                values.append(num)

        if not values:
            logger.warning("No numeric data in Calc!%s — %s is None", col_letter, kpi_key)
            results[kpi_key] = None
            continue

        if agg == "sum":
            results[kpi_key] = sum(values)
        elif agg == "min":
            results[kpi_key] = min(values)
        elif agg == "max":
            results[kpi_key] = max(values)
        else:
            results[kpi_key] = None

    return results


def extract_measures_kpis(wb: Workbook) -> dict[str, float | None]:
    """
    Extract Year 1 totals from the Measures sheet.

    Why: The Measures sheet bridges hourly simulation to annual financials.
    Comparing here lets us isolate aggregation-layer mismatches.

    The Measures sheet has two data regions:
    - Columns C/D: PV+BESS system output and expense/savings
    - Columns F/G: DPPA calculation support and GENCO revenue

    We use a hybrid approach: known cell positions for the most critical
    KPIs, plus label search as fallback for less predictable layouts.
    """
    results: dict[str, float | None] = {}
    sheet_name = "Measures"
    if sheet_name not in wb.sheetnames:
        logger.warning("Sheet '%s' not found — skipping Measures KPIs", sheet_name)
        return results

    ws = wb[sheet_name]
    max_row = ws.max_row or 1

    # --- Label-based search in columns C/D ---
    cd_pairs: list[tuple[str, Any, int]] = []
    for row in range(1, max_row + 1):
        label_cell = ws[f"C{row}"].value
        value_cell = ws[f"D{row}"].value
        if label_cell and isinstance(label_cell, str):
            cd_pairs.append((label_cell.strip().lower(), value_cell, row))

    # --- Label-based search in columns F/G ---
    fg_pairs: list[tuple[str, Any, int]] = []
    for row in range(1, max_row + 1):
        label_cell = ws[f"F{row}"].value
        value_cell = ws[f"G{row}"].value
        if label_cell and isinstance(label_cell, str):
            fg_pairs.append((label_cell.strip().lower(), value_cell, row))

    # Search definitions: (kpi_key, column_pair, search_terms)
    LABEL_SEARCHES: list[tuple[str, str, list[str]]] = [
        ("measures_total_solar_gen", "cd", ["total solar generation"]),
        ("measures_scale_factor", "cd", ["output scale factor", "scale factor"]),
        ("measures_bau_grid_expense", "cd", ["bau grid expense"]),
        ("measures_grid_expense", "cd", ["grid expense", "re grid expense"]),
        ("measures_bess_to_load", "cd", ["bess-to-load", "bess to load"]),
        ("measures_direct_pv_consumption", "cd", ["direct pv consumption"]),
        ("measures_total_dppa_revenue", "fg", ["total revenue"]),
        ("measures_market_energy_payment", "fg", ["marketenergypayment", "market energy payment"]),
        ("measures_cfd_payment", "fg", ["cfdpayment", "cfd payment"]),
    ]

    matched_rows_cd: set[int] = set()
    matched_rows_fg: set[int] = set()

    for kpi_key, col_pair, search_terms in LABEL_SEARCHES:
        pairs = cd_pairs if col_pair == "cd" else fg_pairs
        matched_rows = matched_rows_cd if col_pair == "cd" else matched_rows_fg
        found = False
        for label, value, row in pairs:
            if row in matched_rows:
                continue  # Skip rows already claimed by a more specific match
            if any(term in label for term in search_terms):
                results[kpi_key] = _safe_float(value)
                matched_rows.add(row)
                found = True
                break
        if not found:
            results[kpi_key] = None
            logger.debug("Could not find Measures label for %s", kpi_key)

    # Derive grid savings = BAU - RE grid expense
    bau = results.get("measures_bau_grid_expense")
    re_exp = results.get("measures_grid_expense")
    if bau is not None and re_exp is not None:
        results["measures_total_grid_savings"] = bau - re_exp
    else:
        results["measures_total_grid_savings"] = None

    # First-year revenue from scenario comparison (K6 = DPPA scenario)
    k6_val = ws["K6"].value
    results["measures_first_year_revenue_usd"] = _safe_float(k6_val)

    return results


def extract_all_kpis(excel_path: Path) -> dict[str, float | None]:
    """
    Extract all reference KPIs from a single Excel file.

    Why: This is the main entry point that combines financial, physics,
    and aggregation KPIs into a single dict for JSON serialization.
    """
    logger.info("Loading %s (data_only=True)...", excel_path.name)
    wb = load_workbook(str(excel_path), data_only=True, read_only=False)

    kpis: dict[str, float | None] = {}
    kpis["_source_file"] = None  # placeholder; we'll set it as string in JSON
    kpis.update(extract_financial_kpis(wb))
    kpis.update(extract_calc_column_stats(wb))
    kpis.update(extract_measures_kpis(wb))

    wb.close()

    # Count warnings
    none_count = sum(1 for v in kpis.values() if v is None and not str(v).startswith("_"))
    total_count = sum(1 for k in kpis if not k.startswith("_"))
    if none_count > 0:
        logger.warning(
            "%d of %d KPIs are None for %s — Excel may need recalculation",
            none_count,
            total_count,
            excel_path.name,
        )

    return kpis


def write_reference_json(
    kpis: dict[str, float | None],
    source_path: Path,
    output_dir: Path,
) -> Path:
    """
    Write KPI dict to a JSON reference file.

    Why: JSON files serve as the ground truth for pytest regression tests.
    One file per project keeps references isolated and easy to update.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{source_path.stem}.json"

    serializable = {"_source_file": source_path.name}
    for key, value in kpis.items():
        if key.startswith("_"):
            continue
        serializable[key] = value

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(serializable, f, indent=2, default=str)

    logger.info("Wrote %s (%d KPIs)", output_path, len(serializable) - 1)
    return output_path


def main() -> None:
    """CLI entry point for extracting KPIs from one or more Excel files."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    parser = argparse.ArgumentParser(
        description="Extract reference KPIs from Excel financial model files.",
    )
    parser.add_argument(
        "excel_files",
        nargs="+",
        type=Path,
        help="One or more Excel (.xlsx) files to process.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="Directory for JSON output. Default: tests/data/references/",
    )
    args = parser.parse_args()

    # Resolve output directory relative to project root
    project_root = Path(__file__).resolve().parents[1]
    output_dir = args.output_dir or (project_root / "tests" / "data" / "references")

    for excel_path in args.excel_files:
        excel_path = Path(excel_path).resolve()
        if not excel_path.exists():
            logger.error("File not found: %s", excel_path)
            continue
        if not excel_path.suffix.lower() in (".xlsx", ".xlsm"):
            logger.error("Not an Excel file: %s", excel_path)
            continue

        kpis = extract_all_kpis(excel_path)
        write_reference_json(kpis, excel_path, output_dir)

    logger.info("Done. Reference files written to %s", output_dir)


if __name__ == "__main__":
    main()
