"""
Add "Tariff Schedule 2026" sheet to the Ecoplexus 40MW Excel workbook.

This script duplicates the existing "Tariff Schedule" sheet and remaps the
hour→period assignments to match the Vietnam EVN TOU tariff effective
April 22, 2026.  See docs/tariff_schedules/vietnam_tou_2026.md for the
authoritative mapping.

Usage:
    python scripts/add_ecoplexus_tou2026_sheet.py

The workbook is modified in-place.  Run from the repo root.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

WORKBOOK_PATH = Path(
    "tests/data/projects/AUDIT 20251201 40MW Solar ^M BESS Ecoplexus.xlsx"
)
# The Ecoplexus workbook stores its tariff classification in the Calc sheet's
# TimePeriodFlag column rather than in a standalone "Tariff Schedule" sheet.
# This script creates a new standalone sheet using the TOU 2026 mapping so
# that load_tariff_schedule() can find it without touching the Calc sheet.
NEW_SHEET = "Tariff Schedule 2026"

# New TOU 2026 weekday schedule (0–23, whole-hour mapping).
# See docs/tariff_schedules/vietnam_tou_2026.md for derivation.
TOU2026_WEEKDAY: dict[int, str] = {
    0: "off_peak",
    1: "off_peak",
    2: "off_peak",
    3: "off_peak",
    4: "off_peak",
    5: "off_peak",
    6: "standard",
    7: "standard",
    8: "standard",
    9: "standard",
    10: "standard",
    11: "standard",
    12: "standard",
    13: "standard",
    14: "standard",
    15: "standard",
    16: "standard",
    17: "standard",
    18: "peak",
    19: "peak",
    20: "peak",
    21: "peak",
    22: "peak",
    23: "standard",
}



def add_tou2026_sheet(workbook_path: Path = WORKBOOK_PATH) -> None:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(str(workbook_path))

    if NEW_SHEET in wb.sheetnames:
        print(f"Sheet '{NEW_SHEET}' already exists — removing and recreating.")
        del wb[NEW_SHEET]

    # Create new sheet at the end of the workbook
    new_ws = wb.create_sheet(title=NEW_SHEET)

    # Write header
    new_ws.cell(row=1, column=1, value="hour")
    new_ws.cell(row=1, column=2, value="period")

    # Write 24 rows of hour→period mapping
    for hour in range(24):
        new_ws.cell(row=hour + 2, column=1, value=hour)
        new_ws.cell(row=hour + 2, column=2, value=TOU2026_WEEKDAY[hour])

    wb.save(str(workbook_path))
    print(f"Saved '{NEW_SHEET}' to {workbook_path}")

    # Verify (col A = hour, col B = period, both 1-based so index 0 and 1)
    wb2 = load_workbook(str(workbook_path), read_only=True, data_only=True)
    ws2 = wb2[NEW_SHEET]
    written: dict[int, str] = {}
    for row in ws2.iter_rows(min_row=2):
        h_val = row[0].value
        p_val = row[1].value
        if h_val is not None:
            written[int(h_val)] = str(p_val)
    wb2.close()

    mismatches = [
        (h, TOU2026_WEEKDAY[h], written.get(h))
        for h in range(24)
        if written.get(h) != TOU2026_WEEKDAY[h]
    ]
    if mismatches:
        for h, expected, got in mismatches:
            print(f"  MISMATCH hour {h}: expected={expected}, got={got}")
        raise RuntimeError("Verification failed — see mismatches above.")

    print("Verification passed — all 24 hours match TOU2026 schedule.")


if __name__ == "__main__":
    add_tou2026_sheet()
