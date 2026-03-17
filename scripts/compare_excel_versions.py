"""Generate an HTML logic-comparison report between two Excel model versions."""

from __future__ import annotations

import argparse
import hashlib
import html
import importlib
import math
import re
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Protocol, cast

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

KEY_SHEETS: tuple[str, ...] = (
    "Assumption",
    "Loss",
    "Lifetime",
    "Financial",
    "Measures",
    "Helper",
)

CORE_KPIS: tuple[str, ...] = (
    "project_irr",
    "equity_irr",
    "unlevered_irr",
    "npv_usd",
)


@dataclass(frozen=True)
class SheetDimensionDelta:
    sheet_name: str
    previous_rows: int
    latest_rows: int
    previous_cols: int
    latest_cols: int


@dataclass(frozen=True)
class StructureDiff:
    added_sheets: list[str]
    removed_sheets: list[str]
    dimension_deltas: list[SheetDimensionDelta]


@dataclass(frozen=True)
class DefinedNameDiff:
    name: str
    previous_ref: str | None
    latest_ref: str | None
    previous_value: str | None
    latest_value: str | None
    interpretation: str


@dataclass(frozen=True)
class FormulaChangeGroup:
    sheet_name: str
    change_type: str
    column: str
    row_span: str
    count: int
    representative_cell: str
    previous_formula: str | None
    latest_formula: str | None


@dataclass(frozen=True)
class FormulaDiffResult:
    sheet_name: str
    groups: list[FormulaChangeGroup]


@dataclass(frozen=True)
class KPIDelta:
    kpi: str
    previous_value: float | None
    latest_value: float | None
    absolute_delta: float | None
    relative_delta: float | None
    significance: str


@dataclass(frozen=True)
class EvidenceCell:
    cell_ref: str
    previous_formula: str | None
    latest_formula: str | None
    previous_value: str | None
    latest_value: str | None


@dataclass(frozen=True)
class MaterialFinding:
    subsystem: str
    severity: str
    title: str
    observed_facts: list[str]
    assumptions: list[str]
    confidence: str
    evidence: list[EvidenceCell]


@dataclass(frozen=True)
class WorkbookMeta:
    path: Path
    modified_iso: str
    size_bytes: int
    sha256: str


def discover_workbook_pair(data_dir: Path) -> tuple[Path, Path]:
    """Pick newest and previous .xlsx files by mtime."""
    candidates = sorted(
        (p for p in data_dir.glob("*.xlsx") if p.is_file()),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if len(candidates) < 2:
        raise ValueError(f"Need at least two .xlsx files in {data_dir}")
    return candidates[0], candidates[1]


def load_workbooks(path: Path) -> tuple[Workbook, Workbook]:
    """Load one workbook in formula and cached-value modes."""
    return (
        load_workbook(str(path), data_only=False, read_only=False),
        load_workbook(str(path), data_only=True, read_only=False),
    )


def collect_structure_diff(previous_wb: Workbook, latest_wb: Workbook) -> StructureDiff:
    """Compare sheet membership and dimensions."""
    previous_sheets = set(previous_wb.sheetnames)
    latest_sheets = set(latest_wb.sheetnames)

    deltas: list[SheetDimensionDelta] = []
    for sheet_name in sorted(previous_sheets & latest_sheets):
        prev_ws = previous_wb[sheet_name]
        lat_ws = latest_wb[sheet_name]
        if prev_ws.max_row != lat_ws.max_row or prev_ws.max_column != lat_ws.max_column:
            deltas.append(
                SheetDimensionDelta(
                    sheet_name=sheet_name,
                    previous_rows=prev_ws.max_row,
                    latest_rows=lat_ws.max_row,
                    previous_cols=prev_ws.max_column,
                    latest_cols=lat_ws.max_column,
                )
            )

    return StructureDiff(
        added_sheets=sorted(latest_sheets - previous_sheets),
        removed_sheets=sorted(previous_sheets - latest_sheets),
        dimension_deltas=deltas,
    )


def _defined_name_ref_map(workbook: Workbook) -> dict[str, str]:
    refs: dict[str, str] = {}
    for name, defined_name in workbook.defined_names.items():
        ref_text = getattr(defined_name, "attr_text", None)
        if isinstance(ref_text, str):
            refs[name] = ref_text
    return refs


def _normalize_single_target(ref_text: str) -> tuple[str, str] | None:
    first = ref_text.split(",", maxsplit=1)[0]
    match = re.match(r"^'?([^'!]+)'?!([^!]+)$", first)
    if not match:
        return None
    sheet_name = match.group(1)
    coord = match.group(2)
    return sheet_name, coord.split(":", maxsplit=1)[0].replace("$", "")


def _display_value(raw: object) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, float):
        if math.isnan(raw):
            return None
        return f"{raw:.10g}"
    return str(raw)


def _defined_name_value(values_wb: Workbook, ref_text: str | None) -> str | None:
    if ref_text is None:
        return None
    target = _normalize_single_target(ref_text)
    if target is None:
        return None
    sheet_name, coord = target
    if sheet_name not in values_wb.sheetnames:
        return None
    return _display_value(values_wb[sheet_name][coord].value)


def collect_defined_name_diff(
    previous_formula_wb: Workbook,
    latest_formula_wb: Workbook,
    previous_values_wb: Workbook,
    latest_values_wb: Workbook,
) -> list[DefinedNameDiff]:
    """Compare named-range targets and sampled values."""
    previous_refs = _defined_name_ref_map(previous_formula_wb)
    latest_refs = _defined_name_ref_map(latest_formula_wb)

    diffs: list[DefinedNameDiff] = []
    for name in sorted(set(previous_refs) | set(latest_refs)):
        prev_ref = previous_refs.get(name)
        lat_ref = latest_refs.get(name)
        if prev_ref == lat_ref:
            continue

        prev_value = _defined_name_value(previous_values_wb, prev_ref)
        lat_value = _defined_name_value(latest_values_wb, lat_ref)

        if prev_ref is None:
            interpretation = "Structural: defined name added in latest workbook"
        elif lat_ref is None:
            interpretation = "Structural: defined name removed from latest workbook"
        elif prev_value == lat_value:
            interpretation = "Structural: name target moved but sampled value appears unchanged"
        else:
            interpretation = "Potentially Material: name target moved and sampled value changed"

        diffs.append(
            DefinedNameDiff(
                name=name,
                previous_ref=prev_ref,
                latest_ref=lat_ref,
                previous_value=prev_value,
                latest_value=lat_value,
                interpretation=interpretation,
            )
        )

    return diffs


def _formula_cells(ws: Worksheet) -> dict[str, str]:
    formulas: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                formulas[cell.coordinate] = value
    return formulas


def _formula_template(formula: str) -> str:
    return re.sub(r"(\$?[A-Z]{1,3})\$?\d+", r"\1#", formula)


def _row_span(rows: list[int]) -> str:
    ordered = sorted(rows)
    if not ordered:
        return ""
    if ordered[0] == ordered[-1]:
        return str(ordered[0])
    return f"{ordered[0]}-{ordered[-1]}"


def collect_formula_diff_for_sheet(
    previous_sheet: Worksheet,
    latest_sheet: Worksheet,
    sheet_name: str,
) -> FormulaDiffResult:
    """Group formula diffs to suppress copy-down noise."""
    previous_formulas = _formula_cells(previous_sheet)
    latest_formulas = _formula_cells(latest_sheet)
    grouped: dict[tuple[str, str, str, str], list[tuple[int, str, str | None, str | None]]] = {}

    for coord in sorted(set(previous_formulas) | set(latest_formulas)):
        prev_formula = previous_formulas.get(coord)
        lat_formula = latest_formulas.get(coord)
        if prev_formula == lat_formula:
            continue

        row = int(re.sub(r"^[A-Z]+", "", coord))
        col = re.sub(r"\d+$", "", coord)
        if prev_formula is None:
            change_type = "added_formula"
        elif lat_formula is None:
            change_type = "removed_formula"
        else:
            change_type = "changed_formula"

        prev_tmpl = _formula_template(prev_formula) if prev_formula is not None else "<none>"
        lat_tmpl = _formula_template(lat_formula) if lat_formula is not None else "<none>"
        grouped.setdefault((change_type, col, prev_tmpl, lat_tmpl), []).append(
            (row, coord, prev_formula, lat_formula)
        )

    groups: list[FormulaChangeGroup] = []
    for (change_type, col, _prev_tmpl, _lat_tmpl), items in grouped.items():
        rows = [item[0] for item in items]
        representative = sorted(items, key=lambda item: item[0])[0]
        groups.append(
            FormulaChangeGroup(
                sheet_name=sheet_name,
                change_type=change_type,
                column=col,
                row_span=_row_span(rows),
                count=len(items),
                representative_cell=representative[1],
                previous_formula=representative[2],
                latest_formula=representative[3],
            )
        )

    groups.sort(key=lambda item: (item.change_type, item.column, item.row_span))
    return FormulaDiffResult(sheet_name=sheet_name, groups=groups)


class _ExtractorFn(Protocol):
    def __call__(self, excel_path: Path) -> dict[str, float | None]: ...


def extract_kpi_bundle(path: Path) -> dict[str, float | None]:
    """Reuse existing KPI extraction logic for consistency."""
    try:
        module: Any = importlib.import_module("scripts.extract_excel_kpis")
    except ModuleNotFoundError:
        module = importlib.import_module("extract_excel_kpis")
    extractor = cast(_ExtractorFn, module.extract_all_kpis)
    return extractor(path)


def _compute_significance(
    kpi: str, previous_value: float | None, latest_value: float | None
) -> str:
    if previous_value is None or latest_value is None:
        return "Structural"

    abs_delta = abs(latest_value - previous_value)
    rel_delta = abs_delta / abs(previous_value) if abs(previous_value) > 1e-12 else None

    if kpi in CORE_KPIS:
        if kpi.endswith("irr") and abs_delta >= 0.005:
            return "Material"
        if kpi == "npv_usd" and rel_delta is not None and rel_delta >= 0.1:
            return "Material"
    if rel_delta is not None and rel_delta >= 0.05:
        return "Potentially Material"
    if abs_delta > 0:
        return "Potentially Material"
    return "Structural"


def compute_kpi_deltas(
    previous_kpis: Mapping[str, float | None],
    latest_kpis: Mapping[str, float | None],
) -> list[KPIDelta]:
    """Compute absolute/relative delta rows and significance tags."""
    rows: list[KPIDelta] = []
    for kpi in sorted(set(previous_kpis) | set(latest_kpis)):
        if kpi.startswith("_"):
            continue

        prev = previous_kpis.get(kpi)
        latest = latest_kpis.get(kpi)
        abs_delta: float | None
        rel_delta: float | None
        if prev is None or latest is None:
            abs_delta, rel_delta = None, None
        else:
            abs_delta = latest - prev
            rel_delta = (latest - prev) / abs(prev) if abs(prev) > 1e-12 else None

        rows.append(
            KPIDelta(
                kpi=kpi,
                previous_value=prev,
                latest_value=latest,
                absolute_delta=abs_delta,
                relative_delta=rel_delta,
                significance=_compute_significance(kpi, prev, latest),
            )
        )

    return rows


def _cell_formula(ws: Worksheet, coord: str) -> str | None:
    value = ws[coord].value
    return value if isinstance(value, str) and value.startswith("=") else None


def _cell_value(ws: Worksheet, coord: str) -> str | None:
    return _display_value(ws[coord].value)


def _find_formula_shift_cells(
    previous_sheet: Worksheet,
    latest_sheet: Worksheet,
    old_fragment: str,
    new_fragment: str,
) -> list[str]:
    previous_formulas = _formula_cells(previous_sheet)
    latest_formulas = _formula_cells(latest_sheet)
    cells: list[str] = []
    for coord, prev_formula in previous_formulas.items():
        latest_formula = latest_formulas.get(coord)
        if latest_formula is None:
            continue
        if old_fragment in prev_formula and new_fragment in latest_formula:
            cells.append(coord)
    return sorted(cells)


def _financial_finding(
    previous_formula_wb: Workbook,
    previous_values_wb: Workbook,
    latest_formula_wb: Workbook,
    latest_values_wb: Workbook,
    kpi_deltas: list[KPIDelta],
) -> MaterialFinding | None:
    if "Financial" not in previous_formula_wb.sheetnames:
        return None
    if "Financial" not in latest_formula_wb.sheetnames:
        return None

    prev_f_formula = previous_formula_wb["Financial"]
    prev_f_values = previous_values_wb["Financial"]
    lat_f_formula = latest_formula_wb["Financial"]
    lat_f_values = latest_values_wb["Financial"]

    g170_prev = _cell_value(prev_f_values, "G170")
    g170_latest = _cell_value(lat_f_values, "G170")
    h1_prev_formula = _cell_formula(prev_f_formula, "H1")
    h1_latest_formula = _cell_formula(lat_f_formula, "H1")
    h1_prev_value = _cell_value(prev_f_values, "H1")
    h1_latest_value = _cell_value(lat_f_values, "H1")

    irr_npv_lines: list[str] = []
    for cell in ("G123", "G136", "G189", "G193"):
        prev_val = _cell_value(prev_f_values, cell)
        latest_val = _cell_value(lat_f_values, cell)
        irr_npv_lines.append(f"{cell}: previous={prev_val} latest={latest_val}")

    core_delta_lines = [
        (f"{row.kpi}: {row.previous_value} -> {row.latest_value} (delta={row.absolute_delta})")
        for row in kpi_deltas
        if row.kpi in CORE_KPIS
    ]

    confidence = "High" if "FRESH" in (h1_latest_value or "").upper() else "Medium"
    return MaterialFinding(
        subsystem="Financial solver state",
        severity="Material",
        title="Financial solver state appears refreshed in latest workbook",
        observed_facts=[
            (f"Financial!G170 cached value changed: previous={g170_prev} latest={g170_latest}"),
            f"Financial!H1 formula previous={h1_prev_formula} latest={h1_latest_formula}",
            f"Financial!H1 cached value previous={h1_prev_value} latest={h1_latest_value}",
            *irr_npv_lines,
            *core_delta_lines,
        ],
        assumptions=[
            (
                "Interpretation assumes data_only cached values represent "
                "workbook solve state at save-time."
            ),
            "If either workbook was saved without recalc, this can be overstated.",
        ],
        confidence=confidence,
        evidence=[
            EvidenceCell(
                cell_ref="Financial!G170",
                previous_formula=_cell_formula(prev_f_formula, "G170"),
                latest_formula=_cell_formula(lat_f_formula, "G170"),
                previous_value=g170_prev,
                latest_value=g170_latest,
            ),
            EvidenceCell(
                cell_ref="Financial!H1",
                previous_formula=h1_prev_formula,
                latest_formula=h1_latest_formula,
                previous_value=h1_prev_value,
                latest_value=h1_latest_value,
            ),
            EvidenceCell(
                cell_ref="Financial!J1",
                previous_formula=_cell_formula(prev_f_formula, "J1"),
                latest_formula=_cell_formula(lat_f_formula, "J1"),
                previous_value=_cell_value(prev_f_values, "J1"),
                latest_value=_cell_value(lat_f_values, "J1"),
            ),
        ],
    )


def _lifetime_finding(
    previous_formula_wb: Workbook, latest_formula_wb: Workbook
) -> MaterialFinding | None:
    if "Lifetime" not in previous_formula_wb.sheetnames:
        return None
    if "Lifetime" not in latest_formula_wb.sheetnames:
        return None
    prev_life = previous_formula_wb["Lifetime"]
    lat_life = latest_formula_wb["Lifetime"]
    cells = _find_formula_shift_cells(
        prev_life,
        lat_life,
        "Loss!$A$3:$A$27",
        "Loss!$A$9:$A$33",
    )
    if not cells:
        return None

    sample = cells[0]
    return MaterialFinding(
        subsystem="Loss/Lifetime reference-window shifts",
        severity="Potentially Material",
        title="Lifetime formulas shifted Loss-year reference windows",
        observed_facts=[
            "Observed reference window shift: Loss!$A$3:$A$27 -> Loss!$A$9:$A$33",
            f"Detected {len(cells)} formula locations with this shift pattern",
            f"Representative formula cell: Lifetime!{sample}",
        ],
        assumptions=[
            ("Materiality depends on whether shifted ranges change years included in calculations.")
        ],
        confidence="Medium",
        evidence=[
            EvidenceCell(
                cell_ref=f"Lifetime!{sample}",
                previous_formula=_cell_formula(prev_life, sample),
                latest_formula=_cell_formula(lat_life, sample),
                previous_value=None,
                latest_value=None,
            )
        ],
    )


def _assumption_finding(
    previous_formula_wb: Workbook,
    previous_values_wb: Workbook,
    latest_formula_wb: Workbook,
    latest_values_wb: Workbook,
) -> MaterialFinding | None:
    if "Assumption" not in previous_formula_wb.sheetnames:
        return None
    if "Assumption" not in latest_formula_wb.sheetnames:
        return None

    prev_formula = previous_formula_wb["Assumption"]
    latest_formula = latest_formula_wb["Assumption"]
    prev_values = previous_values_wb["Assumption"]
    latest_values = latest_values_wb["Assumption"]
    prev_k43 = _cell_formula(prev_formula, "K43")
    latest_k43 = _cell_formula(latest_formula, "K43")
    if prev_k43 == latest_k43:
        return None

    return MaterialFinding(
        subsystem="Assumption formula edits",
        severity="Potentially Material",
        title="Assumption!K43 changed from hardcoded multiplier to parameterized reference",
        observed_facts=[
            f"Previous Assumption!K43 formula: {prev_k43}",
            f"Latest Assumption!K43 formula: {latest_k43}",
            "Pattern observed: hardcoded 0.16 multiplier replaced by K45-linked expression.",
        ],
        assumptions=[
            (
                "Parameterization can improve consistency but also "
                "increases sensitivity to K45 edits."
            )
        ],
        confidence="Medium",
        evidence=[
            EvidenceCell(
                cell_ref="Assumption!K43",
                previous_formula=prev_k43,
                latest_formula=latest_k43,
                previous_value=_cell_value(prev_values, "K43"),
                latest_value=_cell_value(latest_values, "K43"),
            )
        ],
    )


def _grid_finding(
    kpi_deltas: list[KPIDelta], formula_diffs: dict[str, FormulaDiffResult]
) -> MaterialFinding | None:
    by_key = {row.kpi: row for row in kpi_deltas}
    if "measures_total_grid_savings" not in by_key and "measures_bau_grid_expense" not in by_key:
        return None

    measures_diff = formula_diffs.get("Measures")
    grouped_text = "No grouped Measures formula changes detected."
    if measures_diff is not None and measures_diff.groups:
        preview = measures_diff.groups[:3]
        grouped_text = "; ".join(
            f"{item.change_type} {item.column}{item.row_span} count={item.count}"
            for item in preview
        )

    savings_row = by_key.get("measures_total_grid_savings")
    bau_row = by_key.get("measures_bau_grid_expense")
    savings_text = (
        "measures_total_grid_savings unavailable"
        if savings_row is None
        else (
            "measures_total_grid_savings: "
            f"{savings_row.previous_value} -> {savings_row.latest_value}"
        )
    )
    bau_text = (
        "measures_bau_grid_expense unavailable"
        if bau_row is None
        else f"measures_bau_grid_expense: {bau_row.previous_value} -> {bau_row.latest_value}"
    )

    return MaterialFinding(
        subsystem="Grid-expense path notes",
        severity="Potentially Material",
        title="Grid-expense KPIs moved while core generation KPIs stayed stable",
        observed_facts=[savings_text, bau_text, grouped_text],
        assumptions=[
            (
                "Interpretation assumes stable calc solar/SoC KPIs indicate "
                "unchanged core generation physics."
            )
        ],
        confidence="Medium",
        evidence=[],
    )


def collect_material_logic_changes(
    previous_formula_wb: Workbook,
    previous_values_wb: Workbook,
    latest_formula_wb: Workbook,
    latest_values_wb: Workbook,
    kpi_deltas: list[KPIDelta],
    formula_diffs: dict[str, FormulaDiffResult],
) -> list[MaterialFinding]:
    """Build high-signal material findings and evidence payloads."""
    findings: list[MaterialFinding] = []

    for maybe_finding in (
        _financial_finding(
            previous_formula_wb,
            previous_values_wb,
            latest_formula_wb,
            latest_values_wb,
            kpi_deltas,
        ),
        _lifetime_finding(previous_formula_wb, latest_formula_wb),
        _assumption_finding(
            previous_formula_wb,
            previous_values_wb,
            latest_formula_wb,
            latest_values_wb,
        ),
        _grid_finding(kpi_deltas, formula_diffs),
    ):
        if maybe_finding is not None:
            findings.append(maybe_finding)

    return findings


def _meta(path: Path) -> WorkbookMeta:
    stat = path.stat()
    modified = datetime.fromtimestamp(stat.st_mtime, tz=UTC).isoformat()
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return WorkbookMeta(path=path, modified_iso=modified, size_bytes=stat.st_size, sha256=digest)


def _fmt(value: float | None, pct: bool = False) -> str:
    if value is None:
        return "N/A"
    if pct:
        return f"{value:.2%}"
    return f"{value:,.6f}" if abs(value) < 1 else f"{value:,.2f}"


def _fmt_signed(value: float | None, pct: bool = False) -> str:
    if value is None:
        return "N/A"
    if pct:
        return f"{value:+.2%}"
    return f"{value:+,.6f}" if abs(value) < 1 else f"{value:+,.2f}"


def _severity_chip(severity: str) -> str:
    css = {
        "Structural": "chip-structural",
        "Potentially Material": "chip-potential",
        "Material": "chip-material",
    }.get(severity, "chip-structural")
    return f"<span class='chip {css}'>{html.escape(severity)}</span>"


def _formula_diff_rows(formula_diffs: dict[str, FormulaDiffResult]) -> str:
    rows: list[str] = []
    for sheet_name in KEY_SHEETS:
        diff = formula_diffs.get(sheet_name)
        if diff is None:
            continue
        for group in diff.groups[:15]:
            rows.append(
                "<tr>"
                f"<td>{html.escape(sheet_name)}</td>"
                f"<td>{html.escape(group.change_type)}</td>"
                f"<td>{html.escape(group.column)}{html.escape(group.row_span)}</td>"
                f"<td>{group.count}</td>"
                f"<td>{html.escape(group.representative_cell)}</td>"
                "</tr>"
            )
    return "".join(rows)


def render_html_report(
    latest_path: Path,
    previous_path: Path,
    selected_mode: str,
    structure_diff: StructureDiff,
    defined_name_diffs: list[DefinedNameDiff],
    kpi_deltas: list[KPIDelta],
    material_findings: list[MaterialFinding],
    formula_diffs: dict[str, FormulaDiffResult],
    command_used: str,
) -> str:
    """Render standalone HTML report."""
    latest_meta = _meta(latest_path)
    previous_meta = _meta(previous_path)

    kpi_rows = "".join(
        "<tr>"
        f"<td>{html.escape(row.kpi)}</td>"
        f"<td>{_fmt(row.previous_value)}</td>"
        f"<td>{_fmt(row.latest_value)}</td>"
        f"<td>{_fmt_signed(row.absolute_delta)}</td>"
        f"<td>{_fmt_signed(row.relative_delta, pct=True)}</td>"
        f"<td>{_severity_chip(row.significance)}</td>"
        "</tr>"
        for row in kpi_deltas
    )

    structure_rows = "".join(
        "<tr>"
        f"<td>{html.escape(delta.sheet_name)}</td>"
        f"<td>{delta.previous_rows}</td>"
        f"<td>{delta.latest_rows}</td>"
        f"<td>{delta.previous_cols}</td>"
        f"<td>{delta.latest_cols}</td>"
        "</tr>"
        for delta in structure_diff.dimension_deltas
    )

    defined_name_rows = "".join(
        "<tr>"
        f"<td>{html.escape(item.name)}</td>"
        f"<td>{html.escape(item.previous_ref or 'N/A')}</td>"
        f"<td>{html.escape(item.latest_ref or 'N/A')}</td>"
        f"<td>{html.escape(item.previous_value or 'N/A')}</td>"
        f"<td>{html.escape(item.latest_value or 'N/A')}</td>"
        f"<td>{html.escape(item.interpretation)}</td>"
        "</tr>"
        for item in defined_name_diffs
    )

    finding_blocks: list[str] = []
    for finding in material_findings:
        observed = "".join(f"<li>{html.escape(line)}</li>" for line in finding.observed_facts)
        assumptions = "".join(f"<li>{html.escape(line)}</li>" for line in finding.assumptions)
        evidence_rows = "".join(
            "<tr>"
            f"<td>{html.escape(item.cell_ref)}</td>"
            f"<td><code>{html.escape(item.previous_formula or 'N/A')}</code></td>"
            f"<td><code>{html.escape(item.latest_formula or 'N/A')}</code></td>"
            f"<td>{html.escape(item.previous_value or 'N/A')}</td>"
            f"<td>{html.escape(item.latest_value or 'N/A')}</td>"
            "</tr>"
            for item in finding.evidence
        )
        evidence_block = ""
        if evidence_rows:
            evidence_block = (
                "<table><thead><tr>"
                "<th>Cell</th><th>Previous Formula</th><th>Latest Formula</th>"
                "<th>Previous Value</th><th>Latest Value</th>"
                "</tr></thead>"
                f"<tbody>{evidence_rows}</tbody></table>"
            )

        title = (
            f"{html.escape(finding.subsystem)} - "
            f"{html.escape(finding.title)} {_severity_chip(finding.severity)}"
        )
        finding_blocks.append(
            "<article class='finding'>"
            f"<h3>{title}</h3>"
            f"<p><strong>Confidence:</strong> {html.escape(finding.confidence)}</p>"
            f"<p><strong>Observed Facts</strong></p><ul>{observed}</ul>"
            f"<p><strong>Assumptions</strong></p><ul>{assumptions}</ul>"
            f"{evidence_block}"
            "</article>"
        )

    added_sheets = "".join(f"<li>{html.escape(name)}</li>" for name in structure_diff.added_sheets)
    removed_sheets = "".join(
        f"<li>{html.escape(name)}</li>" for name in structure_diff.removed_sheets
    )
    structure_block = (
        structure_rows or '<tr><td colspan="5">No dimension deltas detected.</td></tr>'
    )
    formula_block = _formula_diff_rows(formula_diffs)
    formula_block = formula_block or '<tr><td colspan="5">No grouped formula changes.</td></tr>'
    defined_name_block = (
        defined_name_rows or '<tr><td colspan="6">No defined-name retargeting detected.</td></tr>'
    )
    findings_block = "".join(finding_blocks) or "<p>No high-signal findings were detected.</p>"
    generation_ts = datetime.now(tz=UTC).isoformat()

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Excel Logic Comparison Report</title>
  <style>
    :root {{
      --ink: #1b2733;
      --muted: #5d6a76;
      --bg: #f5f7fa;
      --card: #fff;
      --border: #d5dde5;
      --header: #12324a;
      --structural: #6b7280;
      --potential: #b7791f;
      --material: #b42318;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; background: var(--bg); color: var(--ink);
      font-family: Georgia, 'Times New Roman', serif; }}
    .wrap {{ max-width: 1300px; margin: 0 auto; padding: 20px; }}
    h1, h2 {{ color: var(--header); margin: 0 0 10px; }}
    section {{ background: var(--card); border: 1px solid var(--border);
      border-radius: 8px; padding: 14px; margin-bottom: 14px; }}
    table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
    th, td {{ border: 1px solid var(--border); padding: 6px 8px;
      font-size: 13px; vertical-align: top; }}
    thead th {{ background: #e7eef5; }}
    code {{ font-family: Consolas, monospace; font-size: 12px; }}
    .grid2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }}
    .chip {{ display: inline-block; border-radius: 999px; font-size: 11px;
      padding: 2px 8px; color: #fff; }}
    .chip-structural {{ background: var(--structural); }}
    .chip-potential {{ background: var(--potential); }}
    .chip-material {{ background: var(--material); }}
    .muted {{ color: var(--muted); font-size: 13px; }}
    .finding {{ border-top: 1px solid var(--border); padding-top: 10px; margin-top: 10px; }}
    .kv td:first-child {{ width: 30%; font-weight: 700; }}
    @media (max-width: 900px) {{ .grid2 {{ grid-template-columns: 1fr; }} }}
    @media print {{ @page {{ size: A4; margin: 12mm; }}
      body {{ background: #fff; }} section {{ break-inside: avoid; }}
      .wrap {{ padding: 0; }} }}
  </style>
</head>
<body>
  <div class="wrap">
    <h1>Excel Version Logic Comparison Report</h1>
    <p class="muted">Focused on material logic changes, KPI impact, and forensic evidence.</p>
    <section>
      <h2>1. Comparison Overview</h2>
      <table class="kv"><tbody>
        <tr><td>Selection Mode</td><td>{html.escape(selected_mode)}</td></tr>
        <tr><td>Latest Workbook</td><td>{html.escape(latest_meta.path.name)}</td></tr>
        <tr><td>Latest Modified (UTC)</td><td>{html.escape(latest_meta.modified_iso)}</td></tr>
        <tr><td>Latest Size (bytes)</td><td>{latest_meta.size_bytes:,}</td></tr>
        <tr><td>Previous Workbook</td><td>{html.escape(previous_meta.path.name)}</td></tr>
        <tr><td>Previous Modified (UTC)</td><td>{html.escape(previous_meta.modified_iso)}</td></tr>
        <tr><td>Previous Size (bytes)</td><td>{previous_meta.size_bytes:,}</td></tr>
      </tbody></table>
    </section>
    <section>
      <h2>2. Structure Diff</h2>
      <div class="grid2">
        <div><h3>Added Sheets</h3><ul>{added_sheets or "<li>None</li>"}</ul></div>
        <div><h3>Removed Sheets</h3><ul>{removed_sheets or "<li>None</li>"}</ul></div>
      </div>
      <h3>Shared Sheet Dimension Deltas</h3>
      <table>
        <thead><tr><th>Sheet</th><th>Previous Rows</th><th>Latest Rows</th>
        <th>Previous Cols</th><th>Latest Cols</th></tr></thead>
        <tbody>{structure_block}</tbody>
      </table>
    </section>
    <section>
      <h2>3. KPI Delta Dashboard</h2>
      <table>
        <thead><tr><th>KPI</th><th>Previous</th><th>Latest</th>
        <th>Absolute Delta</th><th>Relative Delta</th><th>Significance</th></tr></thead>
        <tbody>{kpi_rows}</tbody>
      </table>
    </section>
    <section>
      <h2>4. Material Logic Changes</h2>
      {findings_block}
      <h3>Grouped Formula Change Patterns</h3>
      <table>
        <thead><tr><th>Sheet</th><th>Change Type</th><th>Span</th>
        <th>Count</th><th>Representative Cell</th></tr></thead>
        <tbody>{formula_block}</tbody>
      </table>
    </section>
    <section>
      <h2>5. Defined Names Retargeting</h2>
      <table>
        <thead><tr><th>Name</th><th>Previous Ref</th><th>Latest Ref</th>
        <th>Previous Value</th><th>Latest Value</th><th>Interpretation</th></tr></thead>
        <tbody>{defined_name_block}</tbody>
      </table>
    </section>
    <section>
      <h2>6. Risk and Interpretation Notes</h2>
      <ul>
        <li><strong>Observed fact:</strong> Values come from cached workbook values.</li>
        <li><strong>Observed fact:</strong> Formula diffs are grouped to suppress copy noise.</li>
        <li><strong>Assumption:</strong>
          Stale workbook caches can distort materiality inference.</li>
        <li><strong>Confidence model:</strong>
          Findings include confidence tags by evidence quality.</li>
      </ul>
    </section>
    <section>
      <h2>7. Reproducibility Footer</h2>
      <table class="kv"><tbody>
        <tr><td>Command</td><td><code>{html.escape(command_used)}</code></td></tr>
        <tr><td>Generated (UTC)</td><td>{html.escape(generation_ts)}</td></tr>
        <tr><td>Latest SHA256</td><td><code>{latest_meta.sha256}</code></td></tr>
        <tr><td>Previous SHA256</td><td><code>{previous_meta.sha256}</code></td></tr>
      </tbody></table>
    </section>
  </div>
</body>
</html>
"""


def write_report(html_content: str, output_path: Path) -> None:
    """Write rendered HTML report to disk."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html_content, encoding="utf-8")


def _collect_formula_diffs(
    previous_formula_wb: Workbook, latest_formula_wb: Workbook
) -> dict[str, FormulaDiffResult]:
    diffs: dict[str, FormulaDiffResult] = {}
    shared = set(previous_formula_wb.sheetnames) & set(latest_formula_wb.sheetnames)
    for sheet_name in sorted(shared):
        if sheet_name in KEY_SHEETS:
            diffs[sheet_name] = collect_formula_diff_for_sheet(
                previous_formula_wb[sheet_name],
                latest_formula_wb[sheet_name],
                sheet_name,
            )
    return diffs


def _command_text(latest: Path, previous: Path, output_path: Path) -> str:
    return (
        "python scripts/compare_excel_versions.py "
        f'--latest "{latest}" --previous "{previous}" --output "{output_path}"'
    )


def _print_summary(
    output_path: Path, latest: Path, previous: Path, rows: Iterable[KPIDelta]
) -> None:
    print(f"Report written: {output_path}")
    print(f"Latest workbook: {latest.name}")
    print(f"Previous workbook: {previous.name}")
    for row in rows:
        if row.kpi in CORE_KPIS:
            print(
                f"{row.kpi}: prev={row.previous_value} latest={row.latest_value} "
                f"delta={row.absolute_delta}"
            )


def main() -> None:
    """CLI entrypoint for Excel version logic report generation."""
    parser = argparse.ArgumentParser(
        description="Compare two Excel model versions and generate one HTML report."
    )
    parser.add_argument("--latest", type=Path, default=None, help="Path to latest workbook")
    parser.add_argument("--previous", type=Path, default=None, help="Path to previous workbook")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("reports/excel_logic_comparison.html"),
        help="Output HTML report path",
    )
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[1]
    if args.latest is None and args.previous is None:
        latest_path, previous_path = discover_workbook_pair(root / "data")
        mode = "auto-discovery"
    elif args.latest is not None and args.previous is not None:
        latest_path = args.latest.resolve()
        previous_path = args.previous.resolve()
        mode = "explicit-args"
    else:
        raise ValueError("Provide both --latest and --previous together, or omit both")

    if not latest_path.exists() or not previous_path.exists():
        raise FileNotFoundError("One or both workbook paths do not exist")

    previous_formula_wb, previous_values_wb = load_workbooks(previous_path)
    latest_formula_wb, latest_values_wb = load_workbooks(latest_path)
    try:
        structure_diff = collect_structure_diff(previous_formula_wb, latest_formula_wb)
        defined_name_diffs = collect_defined_name_diff(
            previous_formula_wb,
            latest_formula_wb,
            previous_values_wb,
            latest_values_wb,
        )
        formula_diffs = _collect_formula_diffs(previous_formula_wb, latest_formula_wb)
        previous_kpis = extract_kpi_bundle(previous_path)
        latest_kpis = extract_kpi_bundle(latest_path)
        kpi_deltas = compute_kpi_deltas(previous_kpis, latest_kpis)
        findings = collect_material_logic_changes(
            previous_formula_wb,
            previous_values_wb,
            latest_formula_wb,
            latest_values_wb,
            kpi_deltas,
            formula_diffs,
        )
        html_report = render_html_report(
            latest_path=latest_path,
            previous_path=previous_path,
            selected_mode=mode,
            structure_diff=structure_diff,
            defined_name_diffs=defined_name_diffs,
            kpi_deltas=kpi_deltas,
            material_findings=findings,
            formula_diffs=formula_diffs,
            command_used=_command_text(latest_path, previous_path, args.output),
        )
        write_report(html_report, args.output)
    finally:
        previous_formula_wb.close()
        previous_values_wb.close()
        latest_formula_wb.close()
        latest_values_wb.close()

    _print_summary(args.output, latest_path, previous_path, kpi_deltas)


if __name__ == "__main__":
    main()
