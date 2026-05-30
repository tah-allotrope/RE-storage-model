"""Integration tests for the DPPA assessment script."""

from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

import openpyxl
import pytest

from re_storage.reporting.excel_writer import (
    create_workbook,
    save_workbook,
    write_assessment_sheet,
    write_assumptions_sheet,
    write_comparison_sheet,
    write_cover_sheet,
    write_sensitivity_sheet,
)


PROJECT_DIR = Path(__file__).resolve().parent.parent / "data" / "projects" / "emivest"


@pytest.fixture
def single_json_project(tmp_path: Path) -> Path:
    """Isolated project dir with exactly one JSON + one CSV (the shared dir has two JSONs)."""
    shutil.copy(PROJECT_DIR / "Emivest.json", tmp_path / "Emivest.json")
    shutil.copy(
        PROJECT_DIR / "Emivest additional data.csv",
        tmp_path / "Emivest additional data.csv",
    )
    return tmp_path


def _sample_kpis() -> dict[str, float]:
    return {
        "project_irr": 0.1234,
        "equity_irr": 0.1567,
        "npv_usd": 1_500_000.0,
        "dscr_min": 1.45,
        "simple_payback_years": 7.5,
        "cash_on_cash_yield": 0.12,
        "year1_solar_generation_mwh": 4260.95,
        "year1_dppa_revenue_usd": 212_394.89,
        "year1_grid_savings_usd": 249_876.35,
        "year1_opex_usd": 48_218.61,
        "year1_ebitda_usd": 414_052.64,
        "debt_amount_usd": 16_120_000.0,
        "exchange_rate_usd_vnd": 25_000.0,
    }


@pytest.mark.slow
class TestDppaAssessmentScript:
    def test_generate_assessment_both_topologies(self):
        """Generate workbook with both topologies — verify two assessment sheets."""
        from scripts.generate_dppa_assessment import generate_assessment

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_dual_topology.xlsx"
            result_path = generate_assessment(
                input_path=PROJECT_DIR,
                project_name="Dual Topology Test",
                output_path=output_path,
                ppa_options=[3],
                topology="both",
            )

            assert result_path.exists()
            wb = openpyxl.load_workbook(str(result_path))
            # Should have sheets for both topologies
            assert any("Onsite" in name for name in wb.sheetnames), f"Expected onsite sheet in {wb.sheetnames}"
            assert any("Offsite" in name for name in wb.sheetnames), f"Expected offsite sheet in {wb.sheetnames}"

    def test_generate_assessment_offsite_topology(self):
        """Generate workbook with offsite topology only."""
        from scripts.generate_dppa_assessment import generate_assessment

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_offsite.xlsx"
            result_path = generate_assessment(
                input_path=PROJECT_DIR,
                project_name="Offsite Test",
                output_path=output_path,
                ppa_options=[3],
                topology="offsite",
            )

            assert result_path.exists()
            wb = openpyxl.load_workbook(str(result_path))
            assert any("Offsite" in name for name in wb.sheetnames)
    def test_generate_assessment_from_json(self):
        """Run the script with the Emivest JSON fixture, verify output .xlsx exists with 5+ sheets."""
        from scripts.generate_dppa_assessment import generate_assessment

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_assessment.xlsx"
            result_path = generate_assessment(
                input_path=PROJECT_DIR,
                project_name="Test DPPA Assessment",
                output_path=output_path,
                ppa_options=[3],  # Only run PPA option 3 to speed up test
                topology="onsite",
            )

            assert result_path.exists()
            assert result_path.suffix == ".xlsx"

            wb = openpyxl.load_workbook(str(result_path))
            assert len(wb.sheetnames) >= 5
            assert "Cover" in wb.sheetnames
            assert "Assumptions" in wb.sheetnames

            # Cover sheet should have non-empty KPI values
            ws_cover = wb["Cover"]
            found_kpi = False
            for row in range(7, 25):
                value = ws_cover.cell(row=row, column=2).value
                if value is not None and value != "N/A":
                    found_kpi = True
                    break
            assert found_kpi, "Cover sheet should have at least one non-N/A KPI value"


    def test_invalid_tariff_mode_raises(self):
        """generate_assessment rejects an unknown tariff_mode."""
        from scripts.generate_dppa_assessment import generate_assessment

        with pytest.raises(ValueError, match="tariff_mode"):
            generate_assessment(
                input_path=PROJECT_DIR,
                project_name="Bad Mode",
                tariff_mode="3-component",
            )

    def test_generate_assessment_two_component(self, single_json_project: Path):
        """2-component workbook surfaces a non-zero demand-charge savings row (onsite)."""
        from scripts.generate_dppa_assessment import generate_assessment

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_two_component.xlsx"
            result_path = generate_assessment(
                input_path=single_json_project,
                project_name="Two-Component Test",
                output_path=output_path,
                ppa_options=[3],
                topology="onsite",
                tariff_mode="2-component",
            )

            assert result_path.exists()
            wb = openpyxl.load_workbook(str(result_path))
            comparison = next(name for name in wb.sheetnames if "Comparison" in name)
            ws = wb[comparison]
            demand_row = None
            for row in range(1, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value or "").startswith("Demand Charge"):
                    demand_row = row
                    break
            assert demand_row is not None, "Comparison sheet missing Demand Charge Savings row"
            assert float(ws.cell(row=demand_row, column=2).value) > 0.0


class TestAssessmentWorkbookIntegration:
    """Faster integration tests that use the Excel writer directly without full pipeline runs."""

    def test_workbook_has_five_sheets(self):
        """Verify that a workbook with all 5 sheet types can be created."""
        wb = create_workbook()
        write_cover_sheet(wb, "Test Project", {}, _sample_kpis())
        write_assessment_sheet(wb, "Assessment", _sample_kpis(), None)
        write_comparison_sheet(wb, {3: _sample_kpis()})
        write_sensitivity_sheet(wb, {})
        write_assumptions_sheet(wb, {"test_key": "test_value"})

        assert len(wb.sheetnames) == 5
        assert "Cover" in wb.sheetnames
        assert "Assessment" in wb.sheetnames
        assert "Comparison" in wb.sheetnames
        assert "Sensitivity" in wb.sheetnames
        assert "Assumptions" in wb.sheetnames

    def test_workbook_saves_and_reopens(self):
        """Verify that a workbook can be saved and reopened with all sheets intact."""
        wb = create_workbook()
        write_cover_sheet(wb, "Test Project", {}, _sample_kpis())
        write_comparison_sheet(wb, {1: _sample_kpis(), 3: _sample_kpis()})

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test.xlsx"
            save_workbook(wb, output_path)

            reopened = openpyxl.load_workbook(str(output_path))
            assert len(reopened.sheetnames) == 2
            assert "Cover" in reopened.sheetnames
            assert "Comparison" in reopened.sheetnames
