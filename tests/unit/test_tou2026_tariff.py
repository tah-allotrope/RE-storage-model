"""
Unit tests for Vietnam TOU 2026 tariff codification (Phase 1).

Covers:
- load_tariff_schedule_from_json: valid TOU2026 JSON, missing block, invalid hours
- SystemAssumptions.tariff_version field round-trip
- Emivest_TOU2026.json fixture validity
- Ecoplexus "Tariff Schedule 2026" Excel sheet validity
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from re_storage.core.exceptions import InputValidationError
from re_storage.core.types import TimePeriod
from re_storage.inputs.json_loader import (
    load_assumptions_from_json,
    load_tariff_schedule_from_json,
)
from re_storage.inputs.schemas import SystemAssumptions

PROJECT_DIR = Path(__file__).resolve().parents[1] / "data" / "projects" / "emivest"
TOU2026_JSON = PROJECT_DIR / "Emivest_TOU2026.json"
ECOPLEXUS_XLSX = (
    Path(__file__).resolve().parents[1]
    / "data"
    / "projects"
    / "AUDIT 20251201 40MW Solar ^M BESS Ecoplexus.xlsx"
)

# Expected TOU 2026 weekday schedule (from docs/tariff_schedules/vietnam_tou_2026.md)
TOU2026_OFF_PEAK = [0, 1, 2, 3, 4, 5]
TOU2026_STANDARD = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 23]
TOU2026_PEAK = [18, 19, 20, 21, 22]


# ---------------------------------------------------------------------------
# load_tariff_schedule_from_json
# ---------------------------------------------------------------------------


def test_load_tariff_schedule_from_json_returns_none_when_absent(tmp_path: Path) -> None:
    """Returns None when JSON has no tariff_schedule block (backward-compatible)."""
    data = {"tariff_schedule": None}
    # "missing key" case: key not present at all
    jf = tmp_path / "no_schedule.json"
    jf.write_text(json.dumps({"project": "test"}))
    result = load_tariff_schedule_from_json(jf)
    assert result is None


def test_load_tariff_schedule_from_json_tou2026_fixture() -> None:
    """Parses Emivest_TOU2026.json and returns the correct 2026 schedule."""
    schedule = load_tariff_schedule_from_json(TOU2026_JSON)
    assert schedule is not None

    assert sorted(schedule[TimePeriod.OFF_PEAK]) == TOU2026_OFF_PEAK
    assert sorted(schedule[TimePeriod.STANDARD]) == TOU2026_STANDARD
    assert sorted(schedule[TimePeriod.PEAK]) == TOU2026_PEAK


def test_load_tariff_schedule_from_json_covers_all_24_hours() -> None:
    """All 24 hours must be accounted for exactly once."""
    schedule = load_tariff_schedule_from_json(TOU2026_JSON)
    assert schedule is not None

    all_hours = sorted(h for hours in schedule.values() for h in hours)
    assert all_hours == list(range(24))


def test_load_tariff_schedule_from_json_peak_window_is_5_hours() -> None:
    """New peak window = exactly 5 hours (17:30–22:30 → hours 18–22)."""
    schedule = load_tariff_schedule_from_json(TOU2026_JSON)
    assert schedule is not None
    assert len(schedule[TimePeriod.PEAK]) == 5


def test_load_tariff_schedule_from_json_off_peak_is_6_hours() -> None:
    """New off-peak window = exactly 6 hours (00:00–06:00 → hours 0–5)."""
    schedule = load_tariff_schedule_from_json(TOU2026_JSON)
    assert schedule is not None
    assert len(schedule[TimePeriod.OFF_PEAK]) == 6


def test_load_tariff_schedule_from_json_solar_hours_are_standard() -> None:
    """Solar generation hours (06:00–17:30 → whole hours 6–17) are all Standard."""
    schedule = load_tariff_schedule_from_json(TOU2026_JSON)
    assert schedule is not None
    solar_hours = list(range(6, 18))  # 6 through 17 inclusive
    for h in solar_hours:
        assert h in schedule[TimePeriod.STANDARD], (
            f"Hour {h} should be Standard (solar window) but is not"
        )


def test_load_tariff_schedule_from_json_invalid_hour_raises(tmp_path: Path) -> None:
    """Hours outside 0–23 must raise InputValidationError."""
    data = {
        "tariff_schedule": {
            "version": "2026",
            "weekday": {
                "off_peak": [0, 1, 2, 3, 4, 5],
                "standard": [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 23],
                "peak": [18, 19, 20, 21, 99],  # 99 is invalid
            },
        }
    }
    jf = tmp_path / "bad_hour.json"
    jf.write_text(json.dumps(data))
    with pytest.raises(InputValidationError, match="outside 0–23"):
        load_tariff_schedule_from_json(jf)


def test_load_tariff_schedule_from_json_duplicate_hour_raises(tmp_path: Path) -> None:
    """Duplicate hours (same hour in two periods) must raise InputValidationError."""
    data = {
        "tariff_schedule": {
            "version": "2026",
            "weekday": {
                "off_peak": [0, 1, 2, 3, 4, 5],
                "standard": [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 22],
                "peak": [18, 19, 20, 21, 22],  # 22 duplicated
            },
        }
    }
    jf = tmp_path / "dup_hour.json"
    jf.write_text(json.dumps(data))
    with pytest.raises(InputValidationError, match="gaps or duplicates"):
        load_tariff_schedule_from_json(jf)


def test_load_tariff_schedule_from_json_missing_weekday_raises(tmp_path: Path) -> None:
    """Missing weekday sub-object must raise InputValidationError."""
    data = {"tariff_schedule": {"version": "2026"}}
    jf = tmp_path / "no_weekday.json"
    jf.write_text(json.dumps(data))
    with pytest.raises(InputValidationError, match="weekday"):
        load_tariff_schedule_from_json(jf)


# ---------------------------------------------------------------------------
# SystemAssumptions.tariff_version
# ---------------------------------------------------------------------------


def test_system_assumptions_tariff_version_default_is_none() -> None:
    """tariff_version defaults to None (optional field)."""
    assumptions = load_assumptions_from_json(PROJECT_DIR / "Emivest.json")
    assert assumptions.tariff_version is None


def test_system_assumptions_tariff_version_populated_from_tou2026() -> None:
    """Emivest_TOU2026.json populates tariff_version='2026'."""
    assumptions = load_assumptions_from_json(TOU2026_JSON)
    assert assumptions.tariff_version == "2026"


def test_system_assumptions_accepts_tariff_version_field() -> None:
    """SystemAssumptions can be constructed with tariff_version without error."""
    base = load_assumptions_from_json(PROJECT_DIR / "Emivest.json")
    updated = base.model_copy(update={"tariff_version": "2026"})
    assert updated.tariff_version == "2026"


# ---------------------------------------------------------------------------
# Ecoplexus Excel — "Tariff Schedule 2026" sheet
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not ECOPLEXUS_XLSX.exists(), reason="Ecoplexus workbook not found")
def test_ecoplexus_tou2026_sheet_exists() -> None:
    """'Tariff Schedule 2026' sheet was added to the Ecoplexus workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(str(ECOPLEXUS_XLSX), read_only=True)
    assert "Tariff Schedule 2026" in wb.sheetnames, (
        "Run scripts/add_ecoplexus_tou2026_sheet.py to create the sheet."
    )
    wb.close()


@pytest.mark.skipif(not ECOPLEXUS_XLSX.exists(), reason="Ecoplexus workbook not found")
def test_ecoplexus_tou2026_sheet_hours_correct() -> None:
    """'Tariff Schedule 2026' sheet contains exactly the TOU2026 hour mapping."""
    from openpyxl import load_workbook

    wb = load_workbook(str(ECOPLEXUS_XLSX), read_only=True, data_only=True)
    ws = wb["Tariff Schedule 2026"]
    written: dict[int, str] = {}
    for row in ws.iter_rows(min_row=2):
        h_val = row[0].value
        p_val = row[1].value
        if h_val is not None:
            written[int(h_val)] = str(p_val)
    wb.close()

    assert sorted(h for h, p in written.items() if p == "off_peak") == TOU2026_OFF_PEAK
    assert sorted(h for h, p in written.items() if p == "standard") == TOU2026_STANDARD
    assert sorted(h for h, p in written.items() if p == "peak") == TOU2026_PEAK


@pytest.mark.skipif(not ECOPLEXUS_XLSX.exists(), reason="Ecoplexus workbook not found")
def test_ecoplexus_tou2026_sheet_covers_all_24_hours() -> None:
    """'Tariff Schedule 2026' sheet has exactly 24 hour entries (0–23)."""
    from openpyxl import load_workbook

    wb = load_workbook(str(ECOPLEXUS_XLSX), read_only=True, data_only=True)
    ws = wb["Tariff Schedule 2026"]
    hours = [int(row[0].value) for row in ws.iter_rows(min_row=2) if row[0].value is not None]
    wb.close()

    assert sorted(hours) == list(range(24))
