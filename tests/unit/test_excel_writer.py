"""Unit tests for the Excel writer module."""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from re_storage.reporting.excel_writer import (
    create_workbook,
    save_workbook,
    write_assessment_sheet,
    write_assumptions_sheet,
    write_comparison_sheet,
    write_cover_sheet,
    write_sensitivity_sheet,
)


def _sample_kpis() -> dict[str, float]:
    return {
        "project_irr": 0.1234,
        "equity_irr": 0.1567,
        "npv_usd": 1_500_000.0,
        "dscr_min": 1.45,
        "simple_payback_years": 7.5,
        "cash_on_cash_yield": 0.12,
        "year1_solar_generation_mwh": 4260.95,
        "year1_dppa_revenue_usd": 212_394.89,
        "year1_grid_savings_usd": 249_876.35,
        "year1_opex_usd": 48_218.61,
        "year1_ebitda_usd": 414_052.64,
        "debt_amount_usd": 16_120_000.0,
    }


def _sample_annual_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "year": list(range(1, 6)),
            "dppa_revenue_usd": [200_000, 210_000, 220_000, 230_000, 240_000],
            "grid_savings_usd": [250_000, 245_000, 240_000, 235_000, 230_000],
            "total_opex_usd": [50_000, 52_000, 54_000, 56_000, 58_000],
            "ebitda_usd": [400_000, 403_000, 406_000, 409_000, 412_000],
            "dscr": [1.5, 1.48, 1.46, 1.44, 1.42],
        }
    )


class TestCreateWorkbook:
    def test_create_workbook_has_no_default_sheet(self):
        wb = create_workbook()
        assert len(wb.sheetnames) == 0

    def test_create_workbook_has_creator(self):
        wb = create_workbook()
        assert wb.properties.creator == "RE-Storage Model"


class TestCoverSheet:
    def test_write_cover_sheet_has_project_name(self):
        wb = create_workbook()
        write_cover_sheet(wb, "Test Project", {}, _sample_kpis())
        ws = wb["Cover"]
        assert "Test Project" in str(ws.cell(row=1, column=1).value)

    def test_write_cover_sheet_has_kpi_table(self):
        wb = create_workbook()
        write_cover_sheet(wb, "Test Project", {}, _sample_kpis())
        ws = wb["Cover"]
        # Header row should be at row 6
        assert ws.cell(row=6, column=1).value == "Metric"
        assert ws.cell(row=6, column=2).value == "Value"

    def test_write_cover_sheet_has_confidentiality_notice(self):
        wb = create_workbook()
        write_cover_sheet(wb, "Test Project", {}, _sample_kpis())
        ws = wb["Cover"]
        assert "CONFIDENTIAL" in str(ws.cell(row=3, column=1).value)

    def test_write_cover_sheet_handles_nan_values(self):
        kpis = _sample_kpis()
        kpis["equity_irr"] = float("nan")
        wb = create_workbook()
        write_cover_sheet(wb, "Test Project", {}, kpis)
        ws = wb["Cover"]
        # Find the row with equity_irr
        for row in range(7, 25):
            if ws.cell(row=row, column=1).value == "Equity IRR":
                assert ws.cell(row=row, column=2).value == "N/A"
                break


class TestAssessmentSheet:
    def test_write_assessment_sheet_has_kpi_and_proforma(self):
        wb = create_workbook()
        write_assessment_sheet(wb, "Assessment", _sample_kpis(), _sample_annual_df())
        ws = wb["Assessment"]
        # KPI header at row 2
        assert ws.cell(row=2, column=1).value == "Metric"
        # Proforma header should exist
        found_proforma = False
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "Annual Proforma":
                found_proforma = True
                break
        assert found_proforma

    def test_write_assessment_sheet_has_vnd_columns(self):
        wb = create_workbook()
        write_assessment_sheet(wb, "Assessment", _sample_kpis(), _sample_annual_df())
        ws = wb["Assessment"]
        # Header row 2 should have USD and VND columns
        assert ws.cell(row=2, column=2).value == "USD"
        assert ws.cell(row=2, column=3).value == "VND"

    def test_write_assessment_sheet_proforma_has_year_column(self):
        wb = create_workbook()
        write_assessment_sheet(wb, "Assessment", _sample_kpis(), _sample_annual_df())
        ws = wb["Assessment"]
        # Find the proforma header row
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "Year":
                # Data rows should follow
                assert ws.cell(row=row + 1, column=1).value == 1
                break


class TestComparisonSheet:
    def test_write_comparison_sheet_has_all_options(self):
        wb = create_workbook()
        results = {1: _sample_kpis(), 2: _sample_kpis(), 3: _sample_kpis(), 4: _sample_kpis()}
        write_comparison_sheet(wb, results)
        ws = wb["Comparison"]
        # Header should have Option labels
        header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert any("Bundled" in str(h) for h in header_row)
        assert any("Separate" in str(h) for h in header_row)
        assert any("DPPA CfD" in str(h) for h in header_row)
        assert any("Fixed EVN" in str(h) for h in header_row)

    def test_write_comparison_sheet_has_vnd_columns(self):
        wb = create_workbook()
        results = {1: _sample_kpis(), 3: _sample_kpis()}
        write_comparison_sheet(wb, results)
        ws = wb["Comparison"]
        header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert any("USD" in str(h) for h in header_row)
        assert any("VND" in str(h) for h in header_row)

    def test_write_comparison_sheet_highlights_best_value(self):
        wb = create_workbook()
        kpis_low = _sample_kpis()
        kpis_high = _sample_kpis()
        kpis_high["project_irr"] = 0.20
        results = {1: kpis_low, 3: kpis_high}
        write_comparison_sheet(wb, results)
        ws = wb["Comparison"]
        # Find the row with Project IRR
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "Project IRR":
                # The best value should have green fill
                best_found = False
                for col in range(2, ws.max_column + 1):
                    fill = ws.cell(row=row, column=col).fill
                    if fill.patternType == "solid" and fill.fgColor and fill.fgColor.rgb:
                        best_found = True
                        break
                assert best_found
                break


class TestSensitivitySheet:
    def test_write_sensitivity_sheet_writes_data_tables(self):
        wb = create_workbook()
        sensitivity = {
            "strike_price": {
                0.05: _sample_kpis(),
                0.06: _sample_kpis(),
                0.07: _sample_kpis(),
            },
            "interest_rate": {
                0.06: _sample_kpis(),
                0.08: _sample_kpis(),
            },
        }
        write_sensitivity_sheet(wb, sensitivity)
        ws = wb["Sensitivity"]
        # Should have at least one section header
        found_section = False
        for row in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and "Sensitivity:" in str(cell_val):
                found_section = True
                break
        assert found_section


class TestAssumptionsSheet:
    def test_write_assumptions_sheet_has_parameters(self):
        wb = create_workbook()
        assumptions = {
            "actual_capacity_kwp": 3221.0,
            "ppa_option": 3,
            "interest_rate_pct": 6.0,
            "strategy_mode": "peak_shaving",
            "custom_param": "test_value",
        }
        write_assumptions_sheet(wb, assumptions)
        ws = wb["Assumptions"]
        # Header at row 1
        assert ws.cell(row=1, column=1).value == "Parameter"
        assert ws.cell(row=1, column=2).value == "Value"
        # Should have data rows
        found_data = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is not None:
                found_data = True
                break
        assert found_data


class TestSaveWorkbook:
    def test_save_workbook_creates_file(self):
        wb = create_workbook()
        write_cover_sheet(wb, "Test Project", {}, _sample_kpis())
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_output.xlsx"
            result = save_workbook(wb, output_path)
            assert result.exists()
            assert result.suffix == ".xlsx"


class TestNumberFormat:
    def test_number_format_irr(self):
        from re_storage.reporting.excel_writer import _apply_number_format

        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=0.1234)
        _apply_number_format(cell, "project_irr")
        assert cell.number_format == "0.00%"

    def test_number_format_usd(self):
        from re_storage.reporting.excel_writer import _apply_number_format

        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=1_500_000)
        _apply_number_format(cell, "npv_usd")
        assert cell.number_format == "#,##0"

    def test_number_format_dscr(self):
        from re_storage.reporting.excel_writer import _apply_number_format

        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=1.45)
        _apply_number_format(cell, "dscr_min")
        assert cell.number_format == "0.00"

    def test_number_format_years(self):
        from re_storage.reporting.excel_writer import _apply_number_format

        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=7.5)
        _apply_number_format(cell, "simple_payback_years")
        assert cell.number_format == "0.0"


class TestCoverSheetVerdict:
    def test_cover_sheet_has_verdict_section(self):
        from re_storage.reporting.assessment import AssessmentVerdict

        wb = create_workbook()
        verdict = AssessmentVerdict(
            overall="GO",
            equity_irr_status="PASS",
            dscr_status="PASS",
            npv_status="PASS",
            payback_status="PASS",
            details=["All metrics pass"],
        )
        write_cover_sheet(wb, "Test Project", {}, _sample_kpis(), verdict=verdict)
        ws = wb["Cover"]
        found_verdict = False
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value and "ASSESSMENT VERDICT" in str(
                ws.cell(row=row, column=1).value
            ):
                found_verdict = True
                break
        assert found_verdict

    def test_cover_sheet_verdict_colors(self):
        from re_storage.reporting.assessment import AssessmentVerdict

        wb = create_workbook()
        verdict = AssessmentVerdict(
            overall="NO-GO",
            equity_irr_status="FAIL",
            dscr_status="PASS",
            npv_status="CAUTION",
            payback_status="PASS",
            details=["IRR too low"],
        )
        write_cover_sheet(wb, "Test Project", {}, _sample_kpis(), verdict=verdict)
        ws = wb["Cover"]
        # Find the overall verdict row
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and "Overall: NO-GO" in str(val):
                fill = ws.cell(row=row, column=1).fill
                assert fill.patternType == "solid"
                break


class TestAssessmentSheetCharts:
    def test_assessment_sheet_accepts_charts_parameter(self):
        wb = create_workbook()
        write_assessment_sheet(
            wb, "Assessment", _sample_kpis(), _sample_annual_df(), charts=[]
        )
        ws = wb["Assessment"]
        assert ws.title == "Assessment"


class TestBranding:
    def test_branding_header_fill_is_green(self):
        wb = create_workbook()
        write_cover_sheet(wb, "Test Project", {}, _sample_kpis())
        ws = wb["Cover"]
        # Header row at row 6
        fill = ws.cell(row=6, column=1).fill
        assert fill.patternType == "solid"
        assert fill.fgColor.rgb == "002E7D32" or str(fill.fgColor.rgb).endswith("2E7D32")
