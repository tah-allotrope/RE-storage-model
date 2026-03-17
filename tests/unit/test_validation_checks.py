"""
Unit tests for validation.checks module.

These tests ensure cross-cutting validation returns warnings for known
violations and raises InputValidationError when required data is missing.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from re_storage.core.exceptions import InputValidationError
from re_storage.inputs.schemas import SystemAssumptions
from re_storage.validation.checks import (
    validate_augmentation_funding,
    validate_degradation_coverage,
    validate_dppa_revenue,
    validate_energy_balance_series,
    validate_financial_solver_freshness,
    validate_full_model,
    validate_soc_bounds_series,
)


def _assumptions(dppa_enabled: bool = True) -> SystemAssumptions:
    return SystemAssumptions(
        simulation_capacity_kwp=100.0,
        actual_capacity_kwp=120.0,
        usable_bess_capacity_kwh=200.0,
        bess_power_rating_kw=100.0,
        charge_efficiency=0.95,
        discharge_efficiency=0.95,
        strategy_mode=1,
        charging_mode=1,
        charge_start_hour=9,
        charge_end_hour=15,
        precharge_target_hour=17,
        precharge_target_soc_kwh=150.0,
        min_direct_pv_share=0.1,
        active_pv2bess_share=0.8,
        demand_reduction_target=0.2,
        strike_price_usd_per_kwh=0.08,
        k_factor=0.98,
        kpp=1.05,
        bess_enabled=True,
        dppa_enabled=dppa_enabled,
    )


def _hourly_results(imbalanced: bool = False, soc_violation: bool = False) -> pd.DataFrame:
    surplus = [10.0, 10.0]
    if imbalanced:
        surplus[0] = 20.0

    soc_kwh = [50.0, 50.0]
    if soc_violation:
        soc_kwh[1] = 250.0

    return pd.DataFrame(
        {
            "solar_gen_kwh": [100.0, 100.0],
            "direct_consumption_kwh": [60.0, 60.0],
            "charged_kwh": [30.0, 30.0],
            "surplus_kwh": surplus,
            "soc_kwh": soc_kwh,
        }
    )


def _lifetime_results(
    dppa_revenue_usd: float = 0.0,
    augmentation_capex_usd: float = 0.0,
    mra_balance_usd: float = 0.0,
) -> pd.DataFrame:
    years = list(range(1, 26))
    dppa = [dppa_revenue_usd] * 25
    aug_capex = [0.0] * 25
    mra = [mra_balance_usd] * 25
    aug_capex[10] = augmentation_capex_usd  # Year 11 (index 10)
    aug_capex[21] = augmentation_capex_usd  # Year 22 (index 21)
    return pd.DataFrame(
        {
            "year": years,
            "dppa_revenue_usd": dppa,
            "augmentation_capex_usd": aug_capex,
            "mra_balance_usd": mra,
        }
    ).set_index("year", drop=False)


def _write_financial_workbook(path: Path, g170_value: float, h1_value: str | None = None) -> Path:
    wb = Workbook()
    ws = wb.create_sheet("Financial")
    if wb.active is not None:
        wb.remove(wb.active)
    ws.title = "Financial"
    ws["G170"] = g170_value
    if h1_value is not None:
        ws["H1"] = h1_value
    wb.save(path)
    wb.close()
    return path


class TestValidateEnergyBalanceSeries:
    def test_returns_warning_on_imbalance(self) -> None:
        warnings = validate_energy_balance_series(_hourly_results(imbalanced=True))
        assert len(warnings) == 1
        assert "Energy balance failed" in warnings[0]

    def test_raises_on_missing_columns(self) -> None:
        with pytest.raises(InputValidationError, match="Missing required columns"):
            validate_energy_balance_series(pd.DataFrame({"solar_gen_kwh": [100.0]}))


class TestValidateSocBoundsSeries:
    def test_returns_warning_on_violation(self) -> None:
        warnings = validate_soc_bounds_series(
            _hourly_results(soc_violation=True), max_capacity_kwh=200.0
        )
        assert len(warnings) == 1
        assert "SoC bounds" in warnings[0]


class TestValidateDppaRevenue:
    def test_warns_when_enabled_but_zero(self) -> None:
        warnings = validate_dppa_revenue(_lifetime_results(dppa_revenue_usd=0.0), dppa_enabled=True)
        assert len(warnings) == 1
        assert "DPPA is enabled" in warnings[0]

    def test_no_warning_when_disabled(self) -> None:
        warnings = validate_dppa_revenue(
            _lifetime_results(dppa_revenue_usd=0.0), dppa_enabled=False
        )
        assert warnings == []


class TestValidateDegradationCoverage:
    def test_warns_on_missing_years(self) -> None:
        table = pd.DataFrame({"year": [1], "pv_factor": [1.0]})
        warnings = validate_degradation_coverage(table, project_years=2)
        assert len(warnings) == 1
        assert "missing" in warnings[0].lower()


class TestValidateAugmentationFunding:
    def test_warns_when_mra_insufficient(self) -> None:
        warnings = validate_augmentation_funding(
            _lifetime_results(augmentation_capex_usd=1000.0, mra_balance_usd=200.0)
        )
        assert len(warnings) == 2
        assert all("augmentation" in w.lower() for w in warnings)


class TestValidateFullModel:
    def test_collects_multiple_warnings(self) -> None:
        hourly_results = _hourly_results(imbalanced=True, soc_violation=True)
        lifetime_results = _lifetime_results(
            dppa_revenue_usd=0.0,
            augmentation_capex_usd=500.0,
            mra_balance_usd=100.0,
        )
        degradation_table = pd.DataFrame({"year": [1], "pv_factor": [1.0]})

        warnings = validate_full_model(
            hourly_results=hourly_results,
            monthly_results=pd.DataFrame(),
            lifetime_results=lifetime_results,
            assumptions=_assumptions(dppa_enabled=True),
            degradation_table=degradation_table,
            project_years=2,
        )

        assert any("Energy balance failed" in warning for warning in warnings)
        assert any("SoC bounds" in warning for warning in warnings)
        assert any("DPPA is enabled" in warning for warning in warnings)
        assert any("Degradation table" in warning for warning in warnings)
        assert any("augmentation" in warning.lower() for warning in warnings)


class TestValidateFinancialSolverFreshness:
    def test_stale_solver_residual_warns(self, tmp_path: Path) -> None:
        workbook = _write_financial_workbook(tmp_path / "stale.xlsx", g170_value=-8373198.66)

        warnings = validate_financial_solver_freshness(str(workbook))

        assert any("G170" in msg for msg in warnings)

    def test_h1_stale_status_warns(self, tmp_path: Path) -> None:
        workbook = _write_financial_workbook(
            tmp_path / "stale_h1.xlsx",
            g170_value=100.0,
            h1_value="STALE - Re-run Solver",
        )

        warnings = validate_financial_solver_freshness(str(workbook))

        assert any("H1" in msg for msg in warnings)

    def test_fresh_solver_has_no_warning(self, tmp_path: Path) -> None:
        workbook = _write_financial_workbook(
            tmp_path / "fresh.xlsx",
            g170_value=0.0,
            h1_value="FRESH",
        )

        warnings = validate_financial_solver_freshness(str(workbook))

        assert warnings == []
