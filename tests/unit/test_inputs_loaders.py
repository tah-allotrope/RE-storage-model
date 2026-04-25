"""
Unit tests for inputs.loaders module.

Tests cover:
1. Assumption sheet loading and validation
2. Hourly data loading and validation
3. Degradation table loading and validation
4. Tariff schedule loading and validation
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

from re_storage.core.exceptions import DegradationTableError, InputValidationError
from re_storage.core.types import HOURS_PER_LEAP_YEAR, HOURS_PER_YEAR, TimePeriod
from re_storage.inputs.loaders import (
    load_assumptions,
    load_assumptions_from_cells,
    load_degradation_table,
    load_financial_params_from_cells,
    load_hourly_data,
    load_tariff_rates_from_cells,
    load_tariff_schedule,
)


def _write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> Path:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)
    return path


def _assumptions_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "simulation_capacity_kwp": 100.0,
                "actual_capacity_kwp": 120.0,
                "usable_bess_capacity_kwh": 200.0,
                "bess_power_rating_kw": 100.0,
                "charge_efficiency": 0.95,
                "discharge_efficiency": 0.95,
                "strategy_mode": 1,
                "charging_mode": 1,
                "charge_start_hour": 9,
                "charge_end_hour": 15,
                "precharge_target_hour": 17,
                "precharge_target_soc_kwh": 150.0,
                "min_direct_pv_share": 0.1,
                "active_pv2bess_share": 0.8,
                "demand_reduction_target": 0.2,
                "strike_price_usd_per_kwh": 0.08,
                "k_factor": 0.98,
                "kpp": 1.05,
                "bess_enabled": True,
                "dppa_enabled": True,
            }
        ]
    )


def _hourly_frame(rows: int) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "datetime": pd.date_range("2025-01-01", periods=rows, freq="h"),
            "simulation_profile_kw": np.full(rows, 100.0),
            "irradiation_wh_m2": np.full(rows, 500.0),
            "load_kw": np.full(rows, 80.0),
            "fmp_usd_per_kwh": np.full(rows, 0.02),
            "cfmp_usd_per_kwh": np.full(rows, 0.03),
        }
    )


def _degradation_frame(years: int) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "year": np.arange(1, years + 1),
            "pv_factor": np.linspace(1.0, 0.9, years),
            "battery_factor_no_replacement": np.linspace(1.0, 0.8, years),
            "battery_factor_with_replacement": np.linspace(1.0, 0.97, years),
        }
    )


def _tariff_frame() -> pd.DataFrame:
    hours = np.arange(0, 24)
    periods = ["off_peak"] * 8 + ["standard"] * 10 + ["peak"] * 6
    return pd.DataFrame({"hour": hours, "period": periods})


class TestLoadAssumptions:
    """Tests for load_assumptions."""

    def test_load_assumptions_success(self, tmp_path: Path) -> None:
        """Valid assumptions sheet should parse into SystemAssumptions."""
        path = _write_excel(tmp_path / "inputs.xlsx", {"Assumption": _assumptions_frame()})
        assumptions = load_assumptions(path)
        assert assumptions.scale_factor == pytest.approx(1.2)

    def test_load_assumptions_missing_field_raises(self, tmp_path: Path) -> None:
        """Missing required columns should raise InputValidationError."""
        frame = _assumptions_frame().drop(columns=["charge_efficiency"])
        path = _write_excel(tmp_path / "inputs.xlsx", {"Assumption": frame})
        with pytest.raises(InputValidationError, match="Missing required assumptions"):
            load_assumptions(path)

    def test_load_assumptions_wrong_row_count(self, tmp_path: Path) -> None:
        """Assumptions sheet must have exactly one row."""
        frame = pd.concat([_assumptions_frame(), _assumptions_frame()], ignore_index=True)
        path = _write_excel(tmp_path / "inputs.xlsx", {"Assumption": frame})
        with pytest.raises(InputValidationError, match="Expected exactly 1 row"):
            load_assumptions(path)


def _write_assumption_label_workbook(path: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook active sheet unavailable")
    ws.title = "Assumption"

    ws["I2"] = "Project Lifetime"
    ws["K2"] = 20
    ws["I3"] = "Base Rate (floating)"
    ws["K3"] = 0.065
    ws["I4"] = "Debt Margin"
    ws["K4"] = 0.02
    ws["I5"] = "Maximum Debt Tenor"
    ws["K5"] = 10
    ws["I6"] = "Target DSCR"
    ws["K6"] = 1.3
    ws["I7"] = "Target Minimum Equity IRR"
    ws["K7"] = 0.10
    ws["I8"] = "Solar"
    ws["K8"] = 750000
    ws["I9"] = "BESS"
    ws["K9"] = 200000
    ws["I10"] = "BOP"
    ws["K10"] = 4843200
    ws["I11"] = "Land acquisition"
    ws["K11"] = 1200000
    ws["I12"] = "Commercial Operation Date"
    ws["K12"] = datetime(2026, 1, 1)
    ws["I13"] = "Maximum Leverage"
    ws["K13"] = 0.7
    ws["I14"] = "USD/VND"
    ws["K14"] = 26000

    ws["O2"] = "Standard"
    ws["Q2"] = 70
    ws["O3"] = "Peak"
    ws["Q3"] = 120
    ws["O4"] = "Off-Peak"
    ws["Q4"] = 45

    wb.save(path)
    wb.close()
    return path


class TestLoadHourlyData:
    """Tests for load_hourly_data."""

    @pytest.mark.parametrize("rows", [HOURS_PER_YEAR, HOURS_PER_LEAP_YEAR])  # type: ignore[untyped-decorator]
    def test_load_hourly_data_accepts_valid_lengths(self, tmp_path: Path, rows: int) -> None:
        """Should accept 8760 and 8784 rows."""
        path = _write_excel(tmp_path / "inputs.xlsx", {"Data Input": _hourly_frame(rows)})
        df = load_hourly_data(path)
        assert len(df) == rows

    def test_load_hourly_data_rejects_invalid_length(self, tmp_path: Path) -> None:
        """Invalid row count should raise InputValidationError."""
        path = _write_excel(tmp_path / "inputs.xlsx", {"Data Input": _hourly_frame(10)})
        with pytest.raises(InputValidationError, match="Expected 8760 or 8784 rows"):
            load_hourly_data(path)

    def test_load_hourly_data_missing_column(self, tmp_path: Path) -> None:
        """Missing required column should raise InputValidationError."""
        frame = _hourly_frame(HOURS_PER_YEAR).drop(columns=["load_kw"])
        path = _write_excel(tmp_path / "inputs.xlsx", {"Data Input": frame})
        with pytest.raises(InputValidationError, match="Missing required hourly columns"):
            load_hourly_data(path)

    def test_load_hourly_data_negative_values(self, tmp_path: Path) -> None:
        """Negative values in non-negative columns should raise."""
        frame = _hourly_frame(HOURS_PER_YEAR)
        frame.loc[0, "load_kw"] = -5.0
        path = _write_excel(tmp_path / "inputs.xlsx", {"Data Input": frame})
        with pytest.raises(InputValidationError, match="contains negative values"):
            load_hourly_data(path)

    def test_load_hourly_data_with_preamble_rows(self, tmp_path: Path) -> None:
        """Loader should detect shifted header rows in new-style Data Input sheets."""
        rows = HOURS_PER_YEAR
        data = _hourly_frame(rows)
        preamble = pd.DataFrame(
            {
                "HOURLY SIMULATION DATA": ["meta1", "meta2"],
                "Unnamed: 1": [None, None],
                "Unnamed: 2": [None, None],
                "Unnamed: 3": [None, None],
                "Unnamed: 4": [None, None],
                "Unnamed: 5": [None, None],
            }
        )
        renamed = data.rename(
            columns={
                "datetime": "DateTime",
                "simulation_profile_kw": "SimulationProfile_kW",
                "irradiation_wh_m2": "Irradiation_W/m2",
                "load_kw": "Load_kW",
                "fmp_usd_per_kwh": "FMP",
                "cfmp_usd_per_kwh": "CFMP",
            }
        )
        combined = pd.concat([preamble, renamed], ignore_index=True)
        path = _write_excel(tmp_path / "inputs.xlsx", {"Data Input": combined})

        df = load_hourly_data(path)
        assert len(df) == rows
        assert "datetime" in df.columns


def test_load_financial_params_from_cells_converts_tax_year_markers_to_durations(
    tmp_path: Path,
) -> None:
    """Workbook tax markers should become duration-style schedule inputs."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook active sheet unavailable")
    ws.title = "Assumption"

    ws["I2"] = "Project Lifetime"
    ws["K2"] = 20
    ws["I3"] = "Base Rate (floating)"
    ws["K3"] = 0.065
    ws["I4"] = "Debt Margin"
    ws["K4"] = 0.02
    ws["I5"] = "Maximum Debt Tenor"
    ws["K5"] = 10
    ws["I6"] = "Target DSCR"
    ws["K6"] = 1.3
    ws["I7"] = "Target Minimum Equity IRR"
    ws["K7"] = 0.10
    ws["I39"] = "Total Cost (fully in place)"
    ws["I40"] = "Land acquisition"
    ws["K40"] = 1_200_000
    ws["I41"] = "Solar"
    ws["K41"] = 750_000
    ws["I42"] = "BESS"
    ws["K42"] = 200_000
    ws["I43"] = "BOP"
    ws["K43"] = 4_843_200
    ws["I44"] = "Depreciation Tenor (Straight line)"
    ws["K44"] = 20
    ws["I48"] = "Maximum Leverage"
    ws["K48"] = 0.7
    ws["I59"] = "Corporate Tax Rate"
    ws["K59"] = 0.2
    ws["I60"] = "Tax Holiday"
    ws["J60"] = 5
    ws["K60"] = 0.0
    ws["I61"] = "First Discount"
    ws["J61"] = 13
    ws["K61"] = 0.05
    ws["I62"] = "Second discount"
    ws["J62"] = 15
    ws["K62"] = 0.1
    ws["I66"] = "Target Minimum Equity IRR"
    ws["K66"] = 0.10

    ws["C2"] = "Actual installation capacity"
    ws["E2"] = 40360
    ws["C3"] = "Total BESS Storage Capacity"
    ws["E3"] = 66000

    ws["O20"] = "PPA Setting Option"
    ws["Q20"] = 3
    ws["O25"] = "Price Escalation"
    ws["Q25"] = 0.05
    ws["O41"] = "Avg. Sun hours Market Price descend"
    ws["Q41"] = -0.05
    ws["O57"] = "Capacity"
    ws["Q57"] = 8056.115384615385

    path = tmp_path / "tax_markers.xlsx"
    wb.save(path)
    wb.close()

    params = load_financial_params_from_cells(path)

    assert params["tax_holiday_years"] == 4
    assert params["first_discount_years"] == 8
    assert params["second_discount_years"] == 2


class TestLoadDegradationTable:
    """Tests for load_degradation_table."""

    def test_load_degradation_table_success(self, tmp_path: Path) -> None:
        """Valid degradation table should load."""
        path = _write_excel(tmp_path / "inputs.xlsx", {"Loss": _degradation_frame(25)})
        df = load_degradation_table(path)
        assert len(df) >= 25

    def test_missing_years_raise(self, tmp_path: Path) -> None:
        """Missing years should raise DegradationTableError."""
        frame = _degradation_frame(25).iloc[:-1]
        path = _write_excel(tmp_path / "inputs.xlsx", {"Loss": frame})
        with pytest.raises(DegradationTableError, match="Missing degradation years"):
            load_degradation_table(path)

    def test_invalid_factor_raises(self, tmp_path: Path) -> None:
        """Out-of-range factor should raise InputValidationError."""
        frame = _degradation_frame(25)
        frame.loc[0, "pv_factor"] = 1.5
        path = _write_excel(tmp_path / "inputs.xlsx", {"Loss": frame})
        with pytest.raises(InputValidationError, match="out of range"):
            load_degradation_table(path)

    def test_load_degradation_table_with_preamble_rows(self, tmp_path: Path) -> None:
        """Loader should detect shifted Loss header rows in new-style sheets."""
        preamble = pd.DataFrame(
            {
                "DEGRADATION & LOSS FACTORS": [
                    "intro",
                    "another intro",
                    None,
                    "sync row",
                    None,
                ]
            }
        )
        loss = pd.DataFrame(
            {
                "Year": np.arange(1, 26),
                "BESS Annual Loss (%)": np.linspace(0.02, 0.03, 25),
                "BESS Cumulative Retention": np.linspace(1.0, 0.8, 25),
                "PV Annual Loss (%)": np.linspace(0.01, 0.02, 25),
                "PV Cumulative Retention": np.linspace(1.0, 0.9, 25),
                "BESS w/ Replacement": np.linspace(1.0, 0.97, 25),
            }
        )
        combined = pd.concat([preamble, loss], ignore_index=True)
        path = _write_excel(tmp_path / "inputs.xlsx", {"Loss": combined})

        df = load_degradation_table(path)
        assert len(df) == 25
        assert {
            "year",
            "pv_factor",
            "battery_factor_no_replacement",
            "battery_factor_with_replacement",
        }.issubset(df.columns)

    def test_load_degradation_table_reconstructs_uncached_formula_columns(
        self, tmp_path: Path
    ) -> None:
        """Loss loader should rebuild retention factors when cached formula cells are blank."""
        from openpyxl import Workbook

        path = tmp_path / "loss_formulas.xlsx"
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Workbook active sheet unavailable")
        ws.title = "Loss"

        ws["A1"] = "Loss Table"
        ws["F1"] = 11
        ws["A2"] = "Year"
        ws["B2"] = "Battery's Loss"
        ws["C2"] = "Battery"
        ws["D2"] = "PV Loss"
        ws["E2"] = "PV"
        ws["F2"] = "Battery wt Replacement"
        ws.append([1, None, 1.0, None, 1.0, 1.0])
        ws.append([2, 0.0255, None, 0.02, None, None])
        ws.append([3, 0.0370, None, 0.0055, None, None])

        wb.save(path)
        wb.close()

        df = load_degradation_table(path, project_years=3)

        assert df["pv_factor"].tolist() == pytest.approx([1.0, 0.98, 0.9745])
        assert df["battery_factor_no_replacement"].tolist() == pytest.approx([1.0, 0.9745, 0.9375])
        assert df["battery_factor_with_replacement"].tolist() == pytest.approx(
            [1.0, 0.9745, 0.9384435]
        )


class TestLoadTariffSchedule:
    """Tests for load_tariff_schedule."""

    def test_load_tariff_schedule_success(self, tmp_path: Path) -> None:
        """Valid tariff schedule should parse into TimePeriod mapping."""
        path = _write_excel(tmp_path / "inputs.xlsx", {"Tariff Schedule": _tariff_frame()})
        schedule = load_tariff_schedule(path)
        assert schedule[TimePeriod.OFF_PEAK]
        assert schedule[TimePeriod.STANDARD]
        assert schedule[TimePeriod.PEAK]

    def test_invalid_period_raises(self, tmp_path: Path) -> None:
        """Unknown period label should raise InputValidationError."""
        frame = _tariff_frame()
        frame.loc[0, "period"] = "super_peak"
        path = _write_excel(tmp_path / "inputs.xlsx", {"Tariff Schedule": frame})
        with pytest.raises(InputValidationError, match="Invalid tariff period"):
            load_tariff_schedule(path)

    def test_invalid_hour_raises(self, tmp_path: Path) -> None:
        """Hour outside 0-23 should raise InputValidationError."""
        frame = _tariff_frame()
        frame.loc[0, "hour"] = 24
        path = _write_excel(tmp_path / "inputs.xlsx", {"Tariff Schedule": frame})
        with pytest.raises(InputValidationError, match="Invalid hour"):
            load_tariff_schedule(path)


class TestLoadTariffAndFinancialFromCells:
    def test_load_tariff_rates_from_cells(self, tmp_path: Path) -> None:
        path = _write_assumption_label_workbook(tmp_path / "assumption.xlsx")

        rates = load_tariff_rates_from_cells(path)

        assert rates[TimePeriod.STANDARD] == pytest.approx(0.07)
        assert rates[TimePeriod.PEAK] == pytest.approx(0.12)
        assert rates[TimePeriod.OFF_PEAK] == pytest.approx(0.045)

    def test_load_financial_params_from_cells(self, tmp_path: Path) -> None:
        path = _write_assumption_label_workbook(tmp_path / "assumption.xlsx")

        params = load_financial_params_from_cells(path)

        assert params["project_years"] == 20
        assert params["interest_rate_pct"] == pytest.approx(8.5)
        assert params["tenor_years"] == 10
        assert params["target_dscr"] == pytest.approx(1.3)
        assert params["initial_capex_usd"] == pytest.approx(6993200.0)
        assert params["discount_rate_pct"] == pytest.approx(10.0)
        assert params["cod_date"] == "2026-01-01"
        assert params["max_leverage_ratio"] == pytest.approx(0.7)
        assert params["exchange_rate_usd_vnd"] == pytest.approx(26000.0)

    def test_load_tariff_rates_from_cells_supports_ca_labels(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        path = tmp_path / "ca_labels.xlsx"
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Workbook active sheet unavailable")
        ws.title = "Assumption"
        ws["I2"] = "USD/VND"
        ws["K2"] = 26000
        ws["O2"] = "Ca_normal"
        ws["Q2"] = 1811
        ws["O3"] = "Ca_peak"
        ws["Q3"] = 3266
        ws["O4"] = "Ca_offpeak"
        ws["Q4"] = 1146
        wb.save(path)
        wb.close()

        rates = load_tariff_rates_from_cells(path)

        assert rates[TimePeriod.STANDARD] == pytest.approx(1811 / 26000)
        assert rates[TimePeriod.PEAK] == pytest.approx(3266 / 26000)
        assert rates[TimePeriod.OFF_PEAK] == pytest.approx(1146 / 26000)

    def test_load_tariff_rates_from_cells_falls_back_to_other_input_table(
        self, tmp_path: Path
    ) -> None:
        from openpyxl import Workbook

        path = tmp_path / "other_input_tariffs.xlsx"
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Workbook active sheet unavailable")
        ws.title = "Assumption"
        ws["O2"] = "Connection Voltage Level"
        ws["Q2"] = 110
        ws["O3"] = "Tariff Structure"
        ws["Q3"] = "2-component"
        ws["O4"] = "Ca_normal"
        ws["Q4"] = None
        ws["O5"] = "Ca_peak"
        ws["Q5"] = None
        ws["O6"] = "Ca_offpeak"
        ws["Q6"] = None
        ws["I2"] = "USD/VND"
        ws["K2"] = 26000

        other = wb.create_sheet("Other Input")
        other["B10"] = "Cp_demand"
        other["C10"] = 209459
        other["D10"] = 235414
        other["E10"] = 0
        other["B11"] = "Ca_normal"
        other["C11"] = 1253
        other["D11"] = 1275
        other["E11"] = 1811
        other["B12"] = "Ca_peak"
        other["C12"] = 2162
        other["D12"] = 2182
        other["E12"] = 3266
        other["B13"] = "Ca_offpeak"
        other["C13"] = 843
        other["D13"] = 859
        other["E13"] = 1146

        wb.save(path)
        wb.close()

        rates = load_tariff_rates_from_cells(path)

        assert rates[TimePeriod.STANDARD] == pytest.approx(1253 / 26000)
        assert rates[TimePeriod.PEAK] == pytest.approx(2162 / 26000)
        assert rates[TimePeriod.OFF_PEAK] == pytest.approx(843 / 26000)


def test_load_assumptions_from_cells_reads_dispatch_flags(tmp_path: Path) -> None:
    """Workbook dispatch toggles should populate the SystemAssumptions flags."""
    from openpyxl import Workbook

    path = tmp_path / "dispatch_flags.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook active sheet unavailable")
    ws.title = "Assumption"

    ws["C2"] = "Simulation Capacity"
    ws["E2"] = 100
    ws["C3"] = "Actual installation capacity"
    ws["E3"] = 200
    ws["C4"] = "Total BESS Storage Capacity"
    ws["E4"] = 500
    ws["C5"] = "Total BESS Power Output"
    ws["E5"] = 250
    ws["C6"] = "DoD"
    ws["E6"] = 0.8
    ws["C7"] = "HalfCycle Efficiency"
    ws["E7"] = 0.95
    ws["C8"] = "Strategy mode"
    ws["E8"] = 1
    ws["C9"] = "Does BESS System include"
    ws["E9"] = 1
    ws["C10"] = "Demand Reduction Target"
    ws["E10"] = 0.2
    ws["C11"] = "PV2BESS Pre-Charge Mode"
    ws["E11"] = 1
    ws["C12"] = "Pre-Charge_StartHour"
    ws["E12"] = 0
    ws["C13"] = "Pre-Charge_EndHour"
    ws["E13"] = 5
    ws["C14"] = "Min PV directly to load"
    ws["E14"] = 0.1
    ws["C15"] = "Pre-Charge Share of PV"
    ws["E15"] = 0.3
    ws["C16"] = "Precharge_TargetSoC_kWh"
    ws["E16"] = 400
    ws["C17"] = "Precharge_TargetHour"
    ws["E17"] = 18
    ws["C18"] = "After Sunset"
    ws["E18"] = 0
    ws["C19"] = "When Needed"
    ws["E19"] = 0
    ws["C20"] = "Peak"
    ws["E20"] = 1
    ws["C21"] = "Optimize mode 1"
    ws["E21"] = 0

    ws["I2"] = "USD/VND"
    ws["K2"] = 26000
    ws["O2"] = "Does model is actived"
    ws["Q2"] = 1
    ws["O3"] = "Strike Price"
    ws["Q3"] = 1800
    ws["O4"] = "k"
    ws["Q4"] = 1.02
    ws["O5"] = "Kpp_22kv"
    ws["Q5"] = 1.027263
    ws["O6"] = "Kpp_110kv"
    ws["Q6"] = 1.008525
    ws["O7"] = "Connection Voltage Level"
    ws["Q7"] = 22

    wb.save(path)
    wb.close()

    assumptions = load_assumptions_from_cells(path)

    assert assumptions.when_needed is False
    assert assumptions.after_sunset is False
    assert assumptions.peak_mode is True
    assert assumptions.optimize_mode is False


def test_load_assumptions_from_cells_derives_bess_totals_from_standard_size(tmp_path: Path) -> None:
    """Blank total BESS cells should fall back to standard size × quantity."""
    from openpyxl import Workbook

    path = tmp_path / "derived_bess_totals.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook active sheet unavailable")
    ws.title = "Assumption"

    ws["C2"] = "Simulation Capacity"
    ws["E2"] = 100
    ws["C3"] = "Actual installation capacity"
    ws["E3"] = 100
    ws["C4"] = "Standard Storage Capacity"
    ws["E4"] = 330
    ws["C5"] = "Standard Power Output"
    ws["E5"] = 100
    ws["C6"] = "System Qunatity"
    ws["E6"] = 200
    ws["C7"] = "Total BESS Storage Capacity"
    ws["E7"] = None
    ws["C8"] = "Total BESS Power Output"
    ws["E8"] = None
    ws["C9"] = "DoD"
    ws["E9"] = 0.85
    ws["C10"] = "HalfCycle Efficiency"
    ws["E10"] = 0.95
    ws["C11"] = "Strategy mode"
    ws["E11"] = 1
    ws["C12"] = "Does BESS System include"
    ws["E12"] = 1
    ws["C13"] = "PV2BESS Pre-Charge Mode"
    ws["E13"] = 0

    ws["I2"] = "USD/VND"
    ws["K2"] = 26000
    ws["O2"] = "Does model is actived"
    ws["Q2"] = 1
    ws["O3"] = "Strike Price"
    ws["Q3"] = 1800
    ws["O4"] = "k"
    ws["Q4"] = 1.02
    ws["O5"] = "Kpp_22kv"
    ws["Q5"] = 1.027263
    ws["O6"] = "Kpp_110kv"
    ws["Q6"] = 1.008525
    ws["O7"] = "Connection Voltage Level"
    ws["Q7"] = 110

    wb.save(path)
    wb.close()

    assumptions = load_assumptions_from_cells(path)

    assert assumptions.usable_bess_capacity_kwh == pytest.approx(330 * 200 * 0.85)
    assert assumptions.bess_power_rating_kw == pytest.approx(100 * 200)
