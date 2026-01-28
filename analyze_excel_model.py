"""
Excel Financial Model Analyzer
Generates a detailed Markdown research document from complex Excel models.
"""

import os
import re
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import glob


def find_excel_file():
    """Find Excel file in current directory."""
    xlsx_files = glob.glob("*.xlsx")
    if not xlsx_files:
        raise FileNotFoundError("No Excel files found in current directory")
    return xlsx_files[0]


def parse_formula_references(formula):
    """Extract cell and sheet references from a formula."""
    if not formula or not isinstance(formula, str):
        return [], []
    
    # Pattern for sheet references like 'Sheet Name'!A1 or SheetName!A1
    sheet_ref_pattern = r"'?([^'!]+)'?!([A-Z]+\d+(?::[A-Z]+\d+)?)"
    # Pattern for local cell references like A1, $A$1, A1:B10
    cell_ref_pattern = r"(?<![A-Z])(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)"
    
    sheet_refs = re.findall(sheet_ref_pattern, formula)
    local_refs = re.findall(cell_ref_pattern, formula)
    
    return sheet_refs, local_refs


def identify_formula_type(formula):
    """Identify the type/purpose of a formula."""
    if not formula or not isinstance(formula, str):
        return "hardcoded"
    
    formula_upper = formula.upper()
    
    if "XIRR" in formula_upper or "IRR" in formula_upper:
        return "IRR Calculation"
    elif "XNPV" in formula_upper or "NPV" in formula_upper:
        return "NPV Calculation"
    elif "SUM" in formula_upper:
        return "Summation"
    elif "AVERAGE" in formula_upper or "AVG" in formula_upper:
        return "Average"
    elif "MAX" in formula_upper:
        return "Maximum"
    elif "MIN" in formula_upper:
        return "Minimum"
    elif "IF" in formula_upper:
        return "Conditional Logic"
    elif "VLOOKUP" in formula_upper or "HLOOKUP" in formula_upper or "INDEX" in formula_upper:
        return "Lookup"
    elif "PMT" in formula_upper or "PPMT" in formula_upper or "IPMT" in formula_upper:
        return "Loan/Debt Calculation"
    elif any(op in formula for op in ["+", "-", "*", "/"]):
        return "Arithmetic"
    else:
        return "Other Formula"


def translate_formula_to_english(formula, col_headers=None):
    """Translate a formula into readable English."""
    if not formula or not isinstance(formula, str):
        return "Hardcoded value"
    
    formula_upper = formula.upper()
    
    translations = []
    
    if "XIRR" in formula_upper:
        translations.append("Calculates Internal Rate of Return using XIRR (dates-based)")
    elif "IRR" in formula_upper:
        translations.append("Calculates Internal Rate of Return")
    
    if "XNPV" in formula_upper:
        translations.append("Calculates Net Present Value using XNPV (dates-based)")
    elif "NPV" in formula_upper:
        translations.append("Calculates Net Present Value")
    
    if "SUM" in formula_upper:
        translations.append("Sums values")
    
    if "MAX" in formula_upper:
        translations.append("Takes maximum value")
    
    if "MIN" in formula_upper:
        translations.append("Takes minimum value")
    
    if "IF" in formula_upper:
        translations.append("Applies conditional logic")
    
    if "VLOOKUP" in formula_upper:
        translations.append("Looks up value vertically")
    
    if "INDEX" in formula_upper and "MATCH" in formula_upper:
        translations.append("Uses INDEX/MATCH lookup")
    
    if "PMT" in formula_upper:
        translations.append("Calculates payment amount")
    
    if "PPMT" in formula_upper:
        translations.append("Calculates principal payment")
    
    if "IPMT" in formula_upper:
        translations.append("Calculates interest payment")
    
    if "SUMIF" in formula_upper:
        translations.append("Conditional sum")
    
    if "SUMPRODUCT" in formula_upper:
        translations.append("Multiplies arrays and sums result")
    
    if "ROUND" in formula_upper:
        translations.append("Rounds value")
    
    if not translations:
        if any(op in formula for op in ["+", "-", "*", "/"]):
            translations.append("Performs arithmetic calculation")
        else:
            translations.append("Formula-based calculation")
    
    return "; ".join(translations)


def analyze_sheet(ws_formulas, ws_values, sheet_name):
    """Analyze a single sheet and return structured data."""
    analysis = {
        "name": sheet_name,
        "purpose": "",
        "dimensions": (ws_formulas.max_row, ws_formulas.max_column),
        "headers": [],
        "hardcoded_inputs": [],
        "time_series_data": [],
        "calculated_columns": [],
        "formulas": [],
        "cross_sheet_refs": [],
        "financial_calcs": [],
        "anomalies": [],
        "merged_cells": list(ws_formulas.merged_cells.ranges) if ws_formulas.merged_cells else []
    }
    
    max_row = ws_formulas.max_row or 1
    max_col = ws_formulas.max_column or 1
    
    if max_row == 0 or max_col == 0:
        analysis["purpose"] = "Empty sheet"
        return analysis
    
    # Extract headers (first few rows often contain headers)
    headers = {}
    for col in range(1, min(max_col + 1, 200)):
        for row in range(1, min(6, max_row + 1)):  # Check first 5 rows for headers
            cell = ws_values.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 0:
                col_letter = get_column_letter(col)
                if col_letter not in headers:
                    headers[col_letter] = cell.value.strip()[:50]
                break
    
    analysis["headers"] = headers
    
    # Infer purpose from headers
    header_text = " ".join(str(v).lower() for v in headers.values())
    if any(term in header_text for term in ["hour", "8760", "dispatch", "generation"]):
        analysis["purpose"] = "Hourly Dispatch / Time Series Simulation"
    elif any(term in header_text for term in ["cash flow", "cashflow", "revenue", "expense"]):
        analysis["purpose"] = "Financial Cash Flow Projection"
    elif any(term in header_text for term in ["irr", "npv", "return"]):
        analysis["purpose"] = "Financial Returns Analysis"
    elif any(term in header_text for term in ["debt", "loan", "interest", "principal"]):
        analysis["purpose"] = "Debt / Financing Schedule"
    elif any(term in header_text for term in ["solar", "pv", "irradiance", "ghi"]):
        analysis["purpose"] = "Solar Generation Model"
    elif any(term in header_text for term in ["battery", "bess", "storage", "charge", "discharge"]):
        analysis["purpose"] = "Battery Storage Model"
    elif any(term in header_text for term in ["input", "assumption", "parameter"]):
        analysis["purpose"] = "Model Inputs / Assumptions"
    elif any(term in header_text for term in ["summary", "output", "result"]):
        analysis["purpose"] = "Summary / Outputs"
    else:
        analysis["purpose"] = "General Calculations"
    
    # Analyze columns
    col_analysis = defaultdict(lambda: {
        "formulas": 0, "hardcoded": 0, "empty": 0, 
        "sample_formula": None, "sample_value": None,
        "formula_types": set(), "cross_sheet_refs": set()
    })
    
    # Sample rows for analysis (don't scan all 8760 rows for every column)
    sample_rows = list(range(1, min(101, max_row + 1)))  # First 100 rows
    if max_row > 100:
        sample_rows.extend([max_row // 4, max_row // 2, 3 * max_row // 4, max_row])  # Plus key rows
    
    for col in range(1, min(max_col + 1, 200)):
        col_letter = get_column_letter(col)
        
        for row in sample_rows:
            if row > max_row:
                continue
            cell_formula = ws_formulas.cell(row=row, column=col)
            cell_value = ws_values.cell(row=row, column=col)
            
            if cell_formula.value is None and cell_value.value is None:
                col_analysis[col_letter]["empty"] += 1
            elif isinstance(cell_formula.value, str) and cell_formula.value.startswith("="):
                col_analysis[col_letter]["formulas"] += 1
                if col_analysis[col_letter]["sample_formula"] is None:
                    col_analysis[col_letter]["sample_formula"] = cell_formula.value
                    col_analysis[col_letter]["sample_value"] = cell_value.value
                
                formula_type = identify_formula_type(cell_formula.value)
                col_analysis[col_letter]["formula_types"].add(formula_type)
                
                sheet_refs, _ = parse_formula_references(cell_formula.value)
                for ref in sheet_refs:
                    col_analysis[col_letter]["cross_sheet_refs"].add(ref[0])
            else:
                col_analysis[col_letter]["hardcoded"] += 1
                if col_analysis[col_letter]["sample_value"] is None:
                    col_analysis[col_letter]["sample_value"] = cell_value.value
    
    # Categorize columns
    for col_letter, stats in col_analysis.items():
        header = headers.get(col_letter, f"Column {col_letter}")
        total_sampled = stats["formulas"] + stats["hardcoded"]
        
        if total_sampled == 0:
            continue  # Skip empty columns
        
        # Check if it's a time series (many rows of data)
        is_time_series = max_row >= 100 and stats["hardcoded"] > 50
        
        if stats["formulas"] > 0 and stats["hardcoded"] == 0:
            # Pure calculated column
            analysis["calculated_columns"].append({
                "column": col_letter,
                "header": header,
                "formula": stats["sample_formula"],
                "formula_types": list(stats["formula_types"]),
                "sample_value": stats["sample_value"]
            })
        elif stats["formulas"] == 0 and stats["hardcoded"] > 0:
            # Pure hardcoded
            if is_time_series:
                analysis["time_series_data"].append({
                    "column": col_letter,
                    "header": header,
                    "sample_value": stats["sample_value"],
                    "row_count": max_row
                })
            else:
                analysis["hardcoded_inputs"].append({
                    "column": col_letter,
                    "header": header,
                    "sample_value": stats["sample_value"]
                })
        elif stats["formulas"] > 0 and stats["hardcoded"] > 0:
            # Mixed - potential anomaly
            analysis["anomalies"].append({
                "column": col_letter,
                "header": header,
                "issue": f"Mixed formulas ({stats['formulas']}) and hardcoded values ({stats['hardcoded']}) in calculation column",
                "sample_formula": stats["sample_formula"]
            })
        
        # Track cross-sheet references
        if stats["cross_sheet_refs"]:
            analysis["cross_sheet_refs"].append({
                "column": col_letter,
                "references": list(stats["cross_sheet_refs"])
            })
        
        # Track financial calculations
        financial_types = {"IRR Calculation", "NPV Calculation", "Loan/Debt Calculation"}
        found_financial = financial_types.intersection(stats["formula_types"])
        if found_financial:
            analysis["financial_calcs"].append({
                "column": col_letter,
                "header": header,
                "calc_types": list(found_financial),
                "formula": stats["sample_formula"],
                "value": stats["sample_value"]
            })
    
    # Look for specific financial formulas in the entire sheet (scan more thoroughly)
    for row in range(1, min(max_row + 1, 500)):
        for col in range(1, min(max_col + 1, 100)):
            cell = ws_formulas.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_upper = cell.value.upper()
                if any(term in formula_upper for term in ["XIRR", "XNPV", "IRR", "NPV"]):
                    cell_val = ws_values.cell(row=row, column=col)
                    col_letter = get_column_letter(col)
                    # Avoid duplicates
                    existing = [f for f in analysis["financial_calcs"] 
                               if f.get("cell") == f"{col_letter}{row}"]
                    if not existing:
                        analysis["financial_calcs"].append({
                            "cell": f"{col_letter}{row}",
                            "header": headers.get(col_letter, ""),
                            "formula": cell.value,
                            "value": cell_val.value,
                            "calc_types": [identify_formula_type(cell.value)]
                        })
    
    return analysis


def generate_markdown_report(analyses, excel_filename):
    """Generate the markdown report from all sheet analyses."""
    
    md_lines = [
        f"# Excel Financial Model Analysis Report",
        f"",
        f"**Source File:** `{excel_filename}`",
        f"",
        f"**Analysis Date:** Generated by automated analyzer",
        f"",
        f"---",
        f"",
        f"## Executive Summary",
        f"",
        f"This document maps the mathematical relationships and data flows in the Excel financial model.",
        f"",
        f"### Sheets Analyzed:",
        f""
    ]
    
    for analysis in analyses:
        purpose = analysis["purpose"] or "Unknown"
        md_lines.append(f"- **{analysis['name']}**: {purpose}")
    
    md_lines.extend(["", "---", ""])
    
    # Detail for each sheet
    for analysis in analyses:
        md_lines.append(f"## Sheet: {analysis['name']}")
        md_lines.append(f"")
        md_lines.append(f"### Purpose")
        md_lines.append(f"{analysis['purpose']}")
        md_lines.append(f"")
        md_lines.append(f"**Dimensions:** {analysis['dimensions'][0]} rows × {analysis['dimensions'][1]} columns")
        md_lines.append(f"")
        
        # Key Inputs
        md_lines.append(f"### Key Inputs (Hardcoded Values)")
        if analysis["hardcoded_inputs"]:
            md_lines.append(f"")
            for inp in analysis["hardcoded_inputs"][:30]:  # Limit to 30
                val_str = str(inp["sample_value"])[:50] if inp["sample_value"] is not None else "N/A"
                md_lines.append(f"- **{inp['header']}** (Col {inp['column']}): `{val_str}`")
        else:
            md_lines.append(f"No significant hardcoded inputs detected.")
        md_lines.append(f"")
        
        # Time Series Data
        if analysis["time_series_data"]:
            md_lines.append(f"### Time Series Data")
            md_lines.append(f"")
            for ts in analysis["time_series_data"][:20]:
                md_lines.append(f"- **{ts['header']}** (Col {ts['column']}): ~{ts['row_count']} rows of data")
            md_lines.append(f"")
        
        # The Math
        md_lines.append(f"### The Math (Formulas and English Translation)")
        md_lines.append(f"")
        
        if analysis["calculated_columns"]:
            for calc in analysis["calculated_columns"][:40]:  # Limit
                header = calc["header"]
                col = calc["column"]
                formula = calc["formula"] or "N/A"
                english = translate_formula_to_english(formula)
                val = calc.get("sample_value")
                val_str = f" → Sample result: `{val}`" if val is not None else ""
                
                md_lines.append(f"#### Column {col}: {header}")
                md_lines.append(f"")
                md_lines.append(f"- **Formula:** `{formula}`")
                md_lines.append(f"- **English:** {english}{val_str}")
                md_lines.append(f"")
        else:
            md_lines.append(f"No calculated columns detected in sampled rows.")
            md_lines.append(f"")
        
        # Financial Calculations (IRR, NPV)
        if analysis["financial_calcs"]:
            md_lines.append(f"### Financial Calculations (IRR, NPV, Debt)")
            md_lines.append(f"")
            for fc in analysis["financial_calcs"]:
                location = fc.get("cell") or f"Col {fc.get('column')}"
                md_lines.append(f"- **{location}**: `{fc['formula']}`")
                if fc.get("value") is not None:
                    md_lines.append(f"  - **Result:** `{fc['value']}`")
                md_lines.append(f"  - **Type:** {', '.join(fc['calc_types'])}")
            md_lines.append(f"")
        
        # Cross-Sheet Dependencies
        if analysis["cross_sheet_refs"]:
            md_lines.append(f"### Cross-Sheet Dependencies")
            md_lines.append(f"")
            refs_summary = defaultdict(list)
            for ref in analysis["cross_sheet_refs"]:
                for sheet in ref["references"]:
                    refs_summary[sheet].append(ref["column"])
            for sheet, cols in refs_summary.items():
                md_lines.append(f"- References **{sheet}** from columns: {', '.join(cols[:10])}")
            md_lines.append(f"")
        
        # Anomalies
        md_lines.append(f"### Anomalies / Flags")
        md_lines.append(f"")
        if analysis["anomalies"]:
            for anom in analysis["anomalies"]:
                md_lines.append(f"- **⚠️ {anom['header']}** (Col {anom['column']}): {anom['issue']}")
                if anom.get("sample_formula"):
                    md_lines.append(f"  - Sample formula: `{anom['sample_formula']}`")
        else:
            md_lines.append(f"No anomalies detected.")
        md_lines.append(f"")
        
        # Merged cells note
        if analysis["merged_cells"]:
            md_lines.append(f"### Formatting Notes")
            md_lines.append(f"")
            md_lines.append(f"- **Merged cell regions:** {len(analysis['merged_cells'])} found (decorative/headers)")
            md_lines.append(f"")
        
        md_lines.append(f"---")
        md_lines.append(f"")
    
    # Final summary
    md_lines.extend([
        "## Data Flow Summary",
        "",
        "### Sheet Interdependencies",
        ""
    ])
    
    # Build dependency graph
    dep_graph = {}
    for analysis in analyses:
        sheet_name = analysis["name"]
        deps = set()
        for ref in analysis["cross_sheet_refs"]:
            deps.update(ref["references"])
        if deps:
            dep_graph[sheet_name] = deps
    
    if dep_graph:
        for sheet, deps in dep_graph.items():
            md_lines.append(f"- **{sheet}** depends on: {', '.join(deps)}")
    else:
        md_lines.append("No cross-sheet dependencies detected in sampled data.")
    
    md_lines.extend([
        "",
        "---",
        "",
        "*End of Analysis Report*"
    ])
    
    return "\n".join(md_lines)


def main():
    print("=" * 60)
    print("Excel Financial Model Analyzer")
    print("=" * 60)
    
    # Find Excel file
    excel_file = find_excel_file()
    print(f"\n📂 Found Excel file: {excel_file}")
    print(f"   File size: {os.path.getsize(excel_file) / (1024*1024):.2f} MB")
    
    # Load workbook twice: once with formulas, once with calculated values
    print("\n📖 Loading workbook (formulas mode)... This may take a minute for large files.")
    wb_formulas = load_workbook(excel_file, data_only=False, read_only=False)
    
    print("📖 Loading workbook (values mode)...")
    wb_values = load_workbook(excel_file, data_only=True, read_only=False)
    
    print(f"\n📋 Found {len(wb_formulas.sheetnames)} sheets:")
    for name in wb_formulas.sheetnames:
        print(f"   - {name}")
    
    # Analyze each sheet
    analyses = []
    print("\n🔍 Analyzing sheets...")
    
    for sheet_name in wb_formulas.sheetnames:
        print(f"   Processing: {sheet_name}...", end=" ")
        try:
            ws_formulas = wb_formulas[sheet_name]
            ws_values = wb_values[sheet_name]
            analysis = analyze_sheet(ws_formulas, ws_values, sheet_name)
            analyses.append(analysis)
            print(f"✓ ({analysis['dimensions'][0]}×{analysis['dimensions'][1]})")
        except Exception as e:
            print(f"⚠️ Error: {e}")
            analyses.append({
                "name": sheet_name,
                "purpose": f"Error during analysis: {e}",
                "dimensions": (0, 0),
                "headers": [],
                "hardcoded_inputs": [],
                "time_series_data": [],
                "calculated_columns": [],
                "formulas": [],
                "cross_sheet_refs": [],
                "financial_calcs": [],
                "anomalies": [],
                "merged_cells": []
            })
    
    # Generate report
    print("\n📝 Generating markdown report...")
    report = generate_markdown_report(analyses, excel_file)
    
    output_file = "analysis_report.md"
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(report)
    
    print(f"\n✅ Report saved to: {output_file}")
    print(f"   Report size: {len(report)} characters")
    
    # Close workbooks
    wb_formulas.close()
    wb_values.close()
    
    print("\n" + "=" * 60)
    print("Analysis complete!")
    print("=" * 60)


if __name__ == "__main__":
    main()
