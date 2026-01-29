"""
Investigation Script: Lifetime Sheet - besstoload Jump Analysis
Purpose: Identify why besstoload values jump at years 11 and 22
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os

def find_excel_file():
    """Find the Excel file in the current directory."""
    for f in os.listdir('.'):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            return f
    return None

def investigate_besstoload(excel_file):
    """Extract besstoload data from Lifetime sheet and related Loss sheet data."""
    
    print(f"Loading workbook: {excel_file}")
    
    # Load with formulas
    wb_formulas = openpyxl.load_workbook(excel_file, data_only=False, read_only=False)
    # Load with calculated values
    wb_values = openpyxl.load_workbook(excel_file, data_only=True, read_only=False)
    
    results = []
    
    # ========== LIFETIME SHEET ANALYSIS ==========
    results.append("=" * 80)
    results.append("LIFETIME SHEET - BESSTOLOAD INVESTIGATION")
    results.append("=" * 80)
    
    if 'Lifetime' in wb_formulas.sheetnames:
        ws_f = wb_formulas['Lifetime']
        ws_v = wb_values['Lifetime']
        
        # First, map row labels to find besstoload
        results.append("\n### Row Labels in Lifetime Sheet (Column A):")
        row_map = {}
        for row in range(1, ws_f.max_row + 1):
            cell_val = ws_v[f'A{row}'].value
            if cell_val:
                row_map[row] = str(cell_val).strip().lower()
                results.append(f"  Row {row}: {cell_val}")
        
        # Find besstoload row (case-insensitive search)
        besstoload_row = None
        for row, label in row_map.items():
            if 'besstoload' in label or 'bess to load' in label or 'bess_to_load' in label:
                besstoload_row = row
                break
        
        if besstoload_row:
            results.append(f"\n### Found besstoload at Row {besstoload_row}")
        else:
            results.append("\n### besstoload row not found by exact match - showing all rows with data")
        
        # Get column headers (Year numbers in row 1)
        results.append("\n### Column Headers (Row 1):")
        col_headers = {}
        for col in range(1, ws_f.max_column + 1):
            header = ws_v.cell(row=1, column=col).value
            if header:
                col_headers[col] = header
                results.append(f"  Col {get_column_letter(col)}: {header}")
        
        # Extract ALL rows with their values to find besstoload
        results.append("\n### All Rows - Values Across Years:")
        results.append("-" * 80)
        
        for row in range(2, ws_f.max_row + 1):
            row_label = ws_v[f'A{row}'].value
            if row_label:
                results.append(f"\n**Row {row}: {row_label}**")
                
                # Get values for years 1-25 (columns B-Z typically)
                values = []
                for col in range(2, min(ws_f.max_column + 1, 28)):  # B to AA
                    val = ws_v.cell(row=row, column=col).value
                    year_num = col - 1  # Year 1 = Col B, etc.
                    if val is not None:
                        values.append(f"Y{year_num}:{val:.4f}" if isinstance(val, (int, float)) else f"Y{year_num}:{val}")
                
                # Show values in groups of 5
                for i in range(0, len(values), 5):
                    results.append("  " + " | ".join(values[i:i+5]))
                
                # Check for jumps at year 11 and 22
                year_values = {}
                for col in range(2, min(ws_f.max_column + 1, 28)):
                    val = ws_v.cell(row=row, column=col).value
                    if isinstance(val, (int, float)):
                        year_values[col - 1] = val
                
                # Detect jumps
                if 10 in year_values and 11 in year_values:
                    change_10_11 = year_values[11] - year_values[10]
                    pct_change = (change_10_11 / year_values[10] * 100) if year_values[10] != 0 else 0
                    if abs(pct_change) > 5:  # More than 5% change
                        results.append(f"  ⚠️ JUMP Y10→Y11: {year_values[10]:.4f} → {year_values[11]:.4f} ({pct_change:+.2f}%)")
                
                if 21 in year_values and 22 in year_values:
                    change_21_22 = year_values[22] - year_values[21]
                    pct_change = (change_21_22 / year_values[21] * 100) if year_values[21] != 0 else 0
                    if abs(pct_change) > 5:
                        results.append(f"  ⚠️ JUMP Y21→Y22: {year_values[21]:.4f} → {year_values[22]:.4f} ({pct_change:+.2f}%)")
        
        # Now get formulas for all rows
        results.append("\n" + "=" * 80)
        results.append("FORMULA ANALYSIS")
        results.append("=" * 80)
        
        for row in range(2, ws_f.max_row + 1):
            row_label = ws_v[f'A{row}'].value
            if row_label:
                results.append(f"\n**Row {row}: {row_label}**")
                
                # Sample formulas at key years
                for year, col in [(1, 2), (10, 11), (11, 12), (21, 22), (22, 23)]:
                    if col <= ws_f.max_column:
                        formula = ws_f.cell(row=row, column=col).value
                        value = ws_v.cell(row=row, column=col).value
                        if formula and isinstance(formula, str) and formula.startswith('='):
                            results.append(f"  Year {year} (Col {get_column_letter(col)}): {formula}")
                            results.append(f"    → Value: {value}")
                        elif formula:
                            results.append(f"  Year {year} (Col {get_column_letter(col)}): HARDCODED = {formula}")
    
    # ========== LOSS SHEET ANALYSIS ==========
    results.append("\n" + "=" * 80)
    results.append("LOSS SHEET - DEGRADATION FACTORS")
    results.append("=" * 80)
    
    if 'Loss' in wb_formulas.sheetnames:
        ws_loss_v = wb_values['Loss']
        ws_loss_f = wb_formulas['Loss']
        
        results.append("\n### Loss Sheet Structure:")
        
        # Get headers
        results.append("\nColumn Headers (Row 1 or 2):")
        for col in range(1, min(ws_loss_v.max_column + 1, 15)):
            h1 = ws_loss_v.cell(row=1, column=col).value
            h2 = ws_loss_v.cell(row=2, column=col).value
            h3 = ws_loss_v.cell(row=3, column=col).value
            results.append(f"  Col {get_column_letter(col)}: R1={h1}, R2={h2}, R3={h3}")
        
        # The Lifetime formulas reference Loss!$A$3:$A$27 and Loss!$E$3:$E$27
        # So let's extract columns A and E (and nearby) from rows 3-27
        results.append("\n### Loss Data (Rows 3-27) - Columns A through H:")
        
        header_row = []
        for col in range(1, 9):
            h = ws_loss_v.cell(row=2, column=col).value
            header_row.append(str(h) if h else f"Col{get_column_letter(col)}")
        results.append("  " + " | ".join(header_row))
        results.append("  " + "-" * 60)
        
        for row in range(3, min(ws_loss_v.max_row + 1, 28)):
            row_data = []
            for col in range(1, 9):
                val = ws_loss_v.cell(row=row, column=col).value
                if isinstance(val, float):
                    row_data.append(f"{val:.4f}")
                elif val is not None:
                    row_data.append(str(val)[:12])
                else:
                    row_data.append("-")
            results.append(f"  Row {row}: " + " | ".join(row_data))
        
        # Check for jumps in Loss sheet degradation factors
        results.append("\n### Loss Sheet - Degradation Jump Analysis:")
        
        for col in range(2, min(ws_loss_v.max_column + 1, 15)):
            col_header = ws_loss_v.cell(row=2, column=col).value
            
            # Get values for years 10, 11, 21, 22
            values = {}
            for row in range(3, min(ws_loss_v.max_row + 1, 28)):
                year_val = ws_loss_v.cell(row=row, column=1).value
                data_val = ws_loss_v.cell(row=row, column=col).value
                if year_val and isinstance(data_val, (int, float)):
                    values[int(year_val) if isinstance(year_val, (int, float)) else 0] = data_val
            
            # Check for jumps
            if 10 in values and 11 in values and values[10] != 0:
                change = (values[11] - values[10]) / values[10] * 100
                if abs(change) > 2:
                    results.append(f"  {col_header} (Col {get_column_letter(col)}): Y10={values[10]:.4f} → Y11={values[11]:.4f} ({change:+.2f}%)")
            
            if 21 in values and 22 in values and values[21] != 0:
                change = (values[22] - values[21]) / values[21] * 100
                if abs(change) > 2:
                    results.append(f"  {col_header} (Col {get_column_letter(col)}): Y21={values[21]:.4f} → Y22={values[22]:.4f} ({change:+.2f}%)")
    
    # ========== OTHER INPUT - AUGMENTATION SCHEDULE ==========
    results.append("\n" + "=" * 80)
    results.append("OTHER INPUT SHEET - AUGMENTATION SCHEDULE")
    results.append("=" * 80)
    
    if 'Other Input' in wb_formulas.sheetnames:
        ws_other = wb_values['Other Input']
        
        results.append("\n### Searching for Augmentation/MRA related data:")
        
        for row in range(1, min(ws_other.max_row + 1, 100)):
            for col in range(1, min(ws_other.max_column + 1, 10)):
                val = ws_other.cell(row=row, column=col).value
                if val and isinstance(val, str):
                    val_lower = val.lower()
                    if any(kw in val_lower for kw in ['augment', 'replace', 'cycle', 'year 11', 'year 22', 'mra', 'capacity']):
                        # Found relevant row, extract the full row
                        row_data = []
                        for c in range(1, min(ws_other.max_column + 1, 10)):
                            cell_val = ws_other.cell(row=row, column=c).value
                            if cell_val is not None:
                                row_data.append(f"{get_column_letter(c)}:{cell_val}")
                        results.append(f"  Row {row}: " + " | ".join(row_data))
                        break
    
    wb_formulas.close()
    wb_values.close()
    
    return "\n".join(results)

if __name__ == "__main__":
    excel_file = find_excel_file()
    if excel_file:
        report = investigate_besstoload(excel_file)
        
        # Write to file
        output_file = "besstoload_investigation.md"
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write("# BESS-to-Load Jump Investigation Report\n\n")
            f.write(report)
        
        print(f"\n✅ Investigation complete. Report saved to: {output_file}")
        print("\n" + "=" * 60)
        print(report)
    else:
        print("❌ No Excel file found in current directory")
