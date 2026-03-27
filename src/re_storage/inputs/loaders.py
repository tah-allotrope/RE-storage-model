"""
Excel/CSV loaders for RE-Storage inputs.

These functions load raw input sheets and apply validation before
passing data to the physics engine.
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from pydantic import ValidationError

from re_storage.core.exceptions import DegradationTableError, InputValidationError
from re_storage.core.types import (
    HOURS_PER_LEAP_YEAR,
    HOURS_PER_YEAR,
    HourlyTimeSeries,
    TimePeriod,
)
from re_storage.inputs.schemas import SystemAssumptions

logger = logging.getLogger(__name__)

ASSUMPTIONS_SHEET = "Assumption"
DATA_INPUT_SHEET = "Data Input"
LOSS_SHEET = "Loss"
TARIFF_SHEET = "Tariff Schedule"
CALC_SHEET = "Calc"
OTHER_INPUT_SHEET = "Other Input"

REQUIRED_HOURLY_COLUMNS = {
    "datetime",
    "simulation_profile_kw",
    "irradiation_wh_m2",
    "load_kw",
    "fmp_usd_per_kwh",
    "cfmp_usd_per_kwh",
}

REQUIRED_DEGRADATION_COLUMNS = {
    "year",
    "pv_factor",
    "battery_factor_no_replacement",
    "battery_factor_with_replacement",
}

_DATA_INPUT_HEADER_MARKERS = {
    "datetime",
    "simulationprofile_kw",
    "irradiation_w/m2",
    "load_kw",
    "fmp",
    "cfmp",
}

_LOSS_HEADER_MARKERS = {
    "year",
    "bess cumulative retention",
    "pv cumulative retention",
    "bess w/ replacement",
}


def load_assumptions(path: Path) -> SystemAssumptions:
    """
    Load and validate the Assumption sheet.

    Args:
        path: Path to Excel input file.

    Returns:
        SystemAssumptions instance.

    Raises:
        InputValidationError: If sheet is missing or invalid.
    """
    df = _read_sheet(path, ASSUMPTIONS_SHEET)

    if len(df) != 1:
        raise InputValidationError(f"Expected exactly 1 row in {ASSUMPTIONS_SHEET}, got {len(df)}.")

    # Only require fields that have no default value in the schema
    required_fields = {
        name
        for name, field in SystemAssumptions.model_fields.items()
        if field.default is None
        and field.default_factory is None  # type: ignore[misc]
        and not field.is_required() is False
    }
    # Simpler: fields are required when FieldInfo.default is PydanticUndefined
    from pydantic_core import PydanticUndefinedType

    required_fields = {
        name
        for name, field in SystemAssumptions.model_fields.items()
        if isinstance(field.default, PydanticUndefinedType) and field.default_factory is None  # type: ignore[misc]
    }
    missing = _missing_columns(df, required_fields)
    if missing:
        raise InputValidationError(f"Missing required assumptions columns: {sorted(missing)}.")

    data = df.iloc[0].to_dict()

    try:
        return SystemAssumptions(**data)
    except ValidationError as exc:  # pragma: no cover - Pydantic raises ValidationError
        raise InputValidationError(f"Assumptions validation failed: {exc}") from exc


def load_assumptions_from_cells(path: Path) -> SystemAssumptions:
    """
    Load assumptions from the real Excel multi-region Assumption sheet.

    The production Assumption sheet has a complex layout with labels and
    values spread across multiple column pairs (C/E, I/K, O/Q). This
    function reads specific cells by label search rather than expecting
    a flat single-row DataFrame.

    Why: The original load_assumptions expects a preprocessed single-row
    sheet. Real Excel files have 67+ rows with section headers, units,
    and multiple data regions. This function bridges that gap.

    Args:
        path: Path to Excel input file.

    Returns:
        SystemAssumptions instance.

    Raises:
        InputValidationError: If the sheet or required values are missing.
    """
    try:
        wb = load_workbook(str(path), data_only=True, read_only=False)
    except (FileNotFoundError, OSError) as exc:
        raise InputValidationError(f"Failed to open Excel file {path}: {exc}") from exc

    if ASSUMPTIONS_SHEET not in wb.sheetnames:
        wb.close()
        raise InputValidationError(f"Sheet '{ASSUMPTIONS_SHEET}' not found in {path}.")

    ws = wb[ASSUMPTIONS_SHEET]

    # Build label→value indexes for each column pair
    ce_map = _build_label_map(ws, label_col="C", value_col="E")
    ik_map = _build_label_map(ws, label_col="I", value_col="K")
    oq_map = _build_label_map(ws, label_col="O", value_col="Q")

    def _get(mapping: dict[str, Any], key: str, default: Any = None) -> Any:
        """Case-insensitive substring label lookup."""
        key_lower = key.lower()
        for k, v in mapping.items():
            if key_lower in k.lower():
                return v
        return default

    def _get_exact(mapping: dict[str, Any], key: str, default: Any = None) -> Any:
        """Case-insensitive exact label lookup (no substring)."""
        key_lower = key.strip().lower()
        for k, v in mapping.items():
            if k.strip().lower() == key_lower:
                return v
        return default

    def _float(mapping: dict[str, Any], key: str, default: float = 0.0) -> float:
        val = _get(mapping, key)
        if val is None:
            logger.warning("Assumption '%s' not found, using default %.4f", key, default)
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            logger.warning("Assumption '%s' has non-numeric value %r, using default", key, val)
            return default

    def _int_val(mapping: dict[str, Any], key: str, default: int = 0) -> int:
        return int(_float(mapping, key, float(default)))

    def _bool_val(mapping: dict[str, Any], key: str, default: bool = False) -> bool:
        val = _float(mapping, key, 1.0 if default else 0.0)
        return val >= 0.5

    # --- Map real Excel cells to SystemAssumptions fields ---
    # Column C/E: System parameters
    simulation_capacity_kwp = _float(ce_map, "Simulation Capacity")
    actual_capacity_kwp = _float(ce_map, "Actual installation capacity")
    total_bess_kwh = _float(ce_map, "Total BESS Storage Capacity")
    total_bess_kw = _float(ce_map, "Total BESS Power Output")
    dod = _float(ce_map, "DoD", 0.85)
    half_cycle_eff = _float(ce_map, "HalfCycle Efficiency", 0.95)
    strategy_mode = _int_val(ce_map, "Strategy mode", 1)
    bess_enabled = _bool_val(ce_map, "Does BESS System include", True)
    demand_reduction_target = _float(ce_map, "Demand Reduction Target", 0.2)

    # Charging parameters
    # PV2BESS Pre-Charge Mode: 0=Off (use time window), 1=Time Window, 2=Precharge
    pv2bess_mode = _int_val(ce_map, "PV2BESS Pre-Charge Mode", 0)
    if pv2bess_mode == 0:
        # Mode 0 = Off: charge all surplus PV into battery at ANY hour.
        # Excel debug confirms PVActive2BESS_kW=0 and PVCharged = full surplus
        # at hours 8-14 (whenever solar > load). No time window restriction.
        # We set the window to 0-23 to allow charging at any hour.
        charging_mode = 1  # Time window
        charge_start = 0
        charge_end = 23
        min_direct_pv = 1.0  # Load is fully served first; only true surplus charges BESS
        pv2bess_share = 1.0  # All surplus goes to BESS
    elif pv2bess_mode == 1:
        charging_mode = 1
        charge_start = _int_val(ce_map, "Pre-Charge_StartHour", 10)
        charge_end = _int_val(ce_map, "Pre-Charge_EndHour", 16)
        min_direct_pv = _float(ce_map, "Min PV directly to load", 0.1)
        pv2bess_share = _float(ce_map, "Pre-Charge Share of PV", 0.1)
    else:
        charging_mode = 2  # Precharge to target
        charge_start = _int_val(ce_map, "Pre-Charge_StartHour", 10)
        charge_end = _int_val(ce_map, "Pre-Charge_EndHour", 16)
        min_direct_pv = _float(ce_map, "Min PV directly to load", 0.1)
        pv2bess_share = _float(ce_map, "Pre-Charge Share of PV", 0.1)
    precharge_soc = _float(ce_map, "Precharge_TargetSoC_kWh", 0.0)
    precharge_hour = _int_val(ce_map, "Precharge_TargetHour", 17)

    # Column O/Q: DPPA parameters
    # Use exact matching for short/ambiguous labels to avoid false matches.
    # "k" would substring-match "Ca_peak", "Kpp_22kv", etc.
    dppa_enabled = _bool_val(oq_map, "Does model is actived", True)
    strike_price_vnd = _float(oq_map, "Strike Price", 1800.0)
    k_factor_val = _get_exact(oq_map, "k")
    if k_factor_val is None:
        logger.warning("DPPA k-factor not found, using default 1.02")
        k_factor_val = 1.02
    else:
        k_factor_val = float(k_factor_val)
    kpp_22 = _float(oq_map, "Kpp_22kv", 1.027)
    kpp_110 = _float(oq_map, "Kpp_110kv", 1.009)
    connection_voltage = _float(oq_map, "Connection Voltage Level", 22.0)

    # Column I/K: Financial parameters
    exchange_rate = _float(ik_map, "USD/VND", 26000.0)
    if exchange_rate <= 0:
        exchange_rate = 26000.0

    # Derive usable BESS capacity (total × DoD)
    usable_bess_kwh = total_bess_kwh * dod

    # Select Kpp based on connection voltage
    kpp = kpp_110 if connection_voltage >= 100 else kpp_22

    # Convert strike price from VND to USD/kWh
    # Excel stores strike price in VND/kWh; divide by exchange rate for USD/kWh
    strike_price_usd_per_kwh = strike_price_vnd / exchange_rate

    wb.close()

    try:
        return SystemAssumptions(
            simulation_capacity_kwp=simulation_capacity_kwp,
            actual_capacity_kwp=actual_capacity_kwp,
            usable_bess_capacity_kwh=usable_bess_kwh,
            bess_power_rating_kw=total_bess_kw,
            charge_efficiency=half_cycle_eff,
            discharge_efficiency=half_cycle_eff,
            strategy_mode=strategy_mode,
            charging_mode=charging_mode,
            charge_start_hour=charge_start,
            charge_end_hour=charge_end,
            precharge_target_hour=precharge_hour,
            precharge_target_soc_kwh=precharge_soc,
            min_direct_pv_share=min_direct_pv,
            active_pv2bess_share=pv2bess_share,
            demand_reduction_target=demand_reduction_target,
            strike_price_usd_per_kwh=strike_price_usd_per_kwh,
            k_factor=k_factor_val,
            kpp=kpp,
            bess_enabled=bess_enabled,
            dppa_enabled=dppa_enabled,
        )
    except ValidationError as exc:
        raise InputValidationError(f"Assumptions validation failed: {exc}") from exc


def _build_label_map(
    ws: Any,
    label_col: str,
    value_col: str,
    max_row: int | None = None,
) -> dict[str, Any]:
    """
    Build a {label_string: value} dict from a worksheet column pair.

    Why: The Assumption sheet stores parameters as label/value pairs in
    adjacent columns (e.g., C=label, E=value). This function scans all
    rows and returns a lookup dict for downstream field extraction.

    Args:
        ws: openpyxl worksheet.
        label_col: Column letter for labels (e.g., "C").
        value_col: Column letter for values (e.g., "E").
        max_row: Maximum row to scan. Defaults to ws.max_row.

    Returns:
        Dict mapping label strings to their corresponding values.
    """
    if max_row is None:
        max_row = ws.max_row or 1
    result: dict[str, Any] = {}
    for row in range(1, max_row + 1):
        label_cell = ws[f"{label_col}{row}"].value
        value_cell = ws[f"{value_col}{row}"].value
        if label_cell and isinstance(label_cell, str) and label_cell.strip():
            result[label_cell.strip()] = value_cell
    return result


def load_tariff_rates_from_cells(path: Path) -> dict[TimePeriod, float]:
    """Load tariff rates from Assumption!O/Q labels."""
    try:
        wb = load_workbook(str(path), data_only=True, read_only=False)
    except (FileNotFoundError, OSError) as exc:
        raise InputValidationError(f"Failed to open Excel file {path}: {exc}") from exc

    if ASSUMPTIONS_SHEET not in wb.sheetnames:
        wb.close()
        raise InputValidationError(f"Sheet '{ASSUMPTIONS_SHEET}' not found in {path}.")

    ws = wb[ASSUMPTIONS_SHEET]
    oq_map = _build_label_map(ws, label_col="O", value_col="Q")
    ik_map = _build_label_map(ws, label_col="I", value_col="K")
    wb.close()

    def _find_exchange_rate() -> float | None:
        for label, raw_value in ik_map.items():
            if "usd/vnd" in label.strip().lower():
                try:
                    value = float(raw_value)
                except (TypeError, ValueError) as exc:
                    raise InputValidationError(
                        f"Financial label '{label}' has non-numeric value {raw_value!r}."
                    ) from exc
                if value > 0:
                    return value
        return None

    exchange_rate = _find_exchange_rate()

    def _find_value(*label_options: str) -> float:
        normalized_options = tuple(option.strip().lower() for option in label_options)
        for label, raw_value in oq_map.items():
            normalized_label = label.strip().lower()
            if any(option in normalized_label for option in normalized_options):
                try:
                    value = float(raw_value)
                except (TypeError, ValueError) as exc:
                    raise InputValidationError(
                        f"Tariff label '{label}' has non-numeric value {raw_value!r}."
                    ) from exc
                if value < 0:
                    raise InputValidationError(
                        f"Tariff label '{label}' has negative value {value}."
                    )
                return value, normalized_label
        raise InputValidationError(f"Missing tariff label containing one of {label_options!r}.")

    def _normalize(value: float, normalized_label: str) -> float:
        if exchange_rate is not None and value >= 100.0 and "ca_" in normalized_label:
            return value / exchange_rate
        return value / 1000.0 if value > 2.0 else value

    off_peak_value, off_peak_label = _find_value("off-peak", "ca_offpeak")
    standard_value, standard_label = _find_value("standard", "ca_normal")
    peak_value, peak_label = _find_value("peak", "ca_peak")

    return {
        TimePeriod.OFF_PEAK: _normalize(off_peak_value, off_peak_label),
        TimePeriod.STANDARD: _normalize(standard_value, standard_label),
        TimePeriod.PEAK: _normalize(peak_value, peak_label),
    }


def load_financial_params_from_cells(path: Path) -> dict[str, float | int | str]:
    """Load key financial parameters from Assumption!I/K, O/Q, and C/E labels."""
    try:
        wb = load_workbook(str(path), data_only=True, read_only=False)
    except (FileNotFoundError, OSError) as exc:
        raise InputValidationError(f"Failed to open Excel file {path}: {exc}") from exc

    if ASSUMPTIONS_SHEET not in wb.sheetnames:
        wb.close()
        raise InputValidationError(f"Sheet '{ASSUMPTIONS_SHEET}' not found in {path}.")

    ws = wb[ASSUMPTIONS_SHEET]
    ik_map = _build_label_map(ws, label_col="I", value_col="K")
    ij_map = _build_label_map(ws, label_col="I", value_col="J")
    oq_map = _build_label_map(ws, label_col="O", value_col="Q")
    ce_map = _build_label_map(ws, label_col="C", value_col="E")
    ik_rows: list[tuple[int, str, Any]] = []
    for row in range(1, (ws.max_row or 1) + 1):
        label = ws[f"I{row}"].value
        value = ws[f"K{row}"].value
        if isinstance(label, str) and label.strip():
            ik_rows.append((row, label.strip(), value))
    wb.close()

    def _find_float(label_substring: str, default: float | None = None) -> float:
        key = label_substring.lower()
        for label, raw_value in ik_map.items():
            if key in label.strip().lower():
                if raw_value is None:
                    # Empty cell next to a matching label — treat as absent and
                    # fall through to the default rather than crashing.
                    continue
                try:
                    return float(raw_value)
                except (TypeError, ValueError) as exc:
                    raise InputValidationError(
                        f"Financial label '{label}' has non-numeric value {raw_value!r}."
                    ) from exc
        if default is not None:
            return default
        raise InputValidationError(f"Missing financial label containing '{label_substring}'.")

    def _find_date_iso(label_substring: str, default_iso: str) -> str:
        key = label_substring.lower()
        for label, raw_value in ik_map.items():
            if key in label.strip().lower():
                if isinstance(raw_value, datetime):
                    return str(raw_value.date())
                if isinstance(raw_value, date):
                    return str(raw_value)
                if isinstance(raw_value, str) and raw_value.strip():
                    return raw_value.strip()
        return default_iso

    def _find_row(label_substring: str) -> int | None:
        key = label_substring.lower()
        for row, label, _value in ik_rows:
            if key in label.lower():
                return row
        return None

    def _find_float_between_rows(
        target_label: str,
        start_row: int,
        end_row: int | None,
        default: float | None = None,
    ) -> float:
        key = target_label.lower()
        for row, label, raw_value in ik_rows:
            if row <= start_row:
                continue
            if end_row is not None and row >= end_row:
                continue
            if key in label.lower():
                try:
                    return float(raw_value)
                except (TypeError, ValueError) as exc:
                    raise InputValidationError(
                        f"Financial label '{label}' has non-numeric value {raw_value!r}."
                    ) from exc
        if default is not None:
            return default
        raise InputValidationError(f"Missing financial label containing '{target_label}'.")

    project_years = int(_find_float("project lifetime", default=25.0))
    base_rate = _find_float("base rate", default=0.06)
    debt_margin = _find_float("debt margin", default=0.0)
    tenor_years = int(_find_float("maximum debt tenor", default=15.0))
    target_dscr = _find_float("target dscr", default=1.3)
    min_equity_irr_ratio = _find_float("target minimum equity irr", default=0.10)
    max_leverage_ratio = _find_float("maximum leverage", default=1.0)
    exchange_rate_usd_vnd = _find_float("usd/vnd", default=26000.0)

    total_cost_row = _find_row("total cost")
    depreciation_row = _find_row("depreciation tenor")
    if total_cost_row is None:
        solar_capex = _find_float("solar", default=0.0)
        bess_capex = _find_float("bess", default=0.0)
        bop_capex = _find_float("bop", default=0.0)
        land_capex = _find_float("land acquisition", default=0.0)
    else:
        solar_capex = _find_float_between_rows(
            "solar", total_cost_row, depreciation_row, default=0.0
        )
        bess_capex = _find_float_between_rows("bess", total_cost_row, depreciation_row, default=0.0)
        bop_capex = _find_float_between_rows("bop", total_cost_row, depreciation_row, default=0.0)
        land_capex = _find_float_between_rows(
            "land acquisition",
            total_cost_row,
            depreciation_row,
            default=0.0,
        )
    initial_capex_usd = max(solar_capex + bess_capex + bop_capex + land_capex, 0.0)

    # --- CAPEX breakdown (unit rates for OPEX computation) ---
    # K41 = solar $/MWp, K42 = BESS $/MWh — used to back out installed capacity
    solar_usd_per_mwp = _find_float("solar capex", default=0.0)
    if solar_usd_per_mwp <= 0:
        solar_usd_per_mwp = _find_float_between_rows(
            "solar", total_cost_row or 0, depreciation_row, default=750_000.0
        )
    bess_usd_per_mwh = _find_float("bess capex", default=0.0)
    if bess_usd_per_mwh <= 0:
        bess_usd_per_mwh = _find_float_between_rows(
            "bess", total_cost_row or 0, depreciation_row, default=200_000.0
        )

    # Derive installed capacity from per-unit rate (fallback to C/E direct read)
    def _ce_float(key: str, default: float = 0.0) -> float:
        key_lower = key.lower()
        for k, v in ce_map.items():
            if key_lower in k.lower():
                try:
                    return float(v)
                except (TypeError, ValueError):
                    pass
        return default

    installed_pv_kwp = _ce_float("actual installation capacity", default=0.0)
    installed_pv_mwp = (
        installed_pv_kwp / 1000.0
        if installed_pv_kwp > 0
        else (solar_capex / solar_usd_per_mwp if solar_usd_per_mwp > 0 else 0.0)
    )
    total_bess_kwh = _ce_float("Total BESS Storage Capacity", default=0.0)
    bess_mwh = (
        total_bess_kwh / 1000.0
        if total_bess_kwh > 0
        else (bess_capex / bess_usd_per_mwh if bess_usd_per_mwh > 0 else 0.0)
    )

    # The Solar and BESS rows in the Total Cost section store unit rates ($/MWp, $/MWh),
    # not absolute costs. Recompute absolute capex using unit rate × installed capacity
    # now that capacities are known. BOP and Land are already absolute ($).
    if installed_pv_mwp > 0 and solar_usd_per_mwp > 0:
        solar_capex = solar_usd_per_mwp * installed_pv_mwp
    if bess_mwh > 0 and bess_usd_per_mwh > 0:
        bess_capex = bess_usd_per_mwh * bess_mwh
    initial_capex_usd = max(solar_capex + bess_capex + bop_capex + land_capex, 0.0)

    # --- OPEX parameters (Assumption!K26–K34) ---
    om_solar_usd_per_mwp = _find_float("o&m", default=6_000.0)
    if om_solar_usd_per_mwp <= 0:
        om_solar_usd_per_mwp = 6_000.0
    om_bess_usd_per_mwh = _find_float("o&m bess", default=2_000.0)
    if om_bess_usd_per_mwh <= 0:
        om_bess_usd_per_mwh = 2_000.0
    insurance_solar_pct = _find_float("insurance", default=0.0025)
    if insurance_solar_pct > 1.0:
        insurance_solar_pct /= 100.0
    insurance_bess_pct = _find_float("insurance bess", default=0.0025)
    if insurance_bess_pct > 1.0:
        insurance_bess_pct /= 100.0
    other_opex_usd_per_mwp = _find_float("other opex", default=1_000.0)
    asset_mgmt_usd_per_mwp = _find_float("asset management", default=3_000.0)
    land_lease_pct_revenue = _find_float("land lease", default=0.0)
    if land_lease_pct_revenue > 1.0:
        land_lease_pct_revenue /= 100.0
    opex_escalation_pct = _find_float("opex escalation", default=0.04)
    if opex_escalation_pct > 1.0:
        opex_escalation_pct /= 100.0

    # --- Tax parameters (Assumption!K44, K62–K65) ---
    depreciation_tenor_years = int(_find_float("depreciation tenor", default=20.0))
    tax_rate = _find_float("corporate tax rate", default=0.20)
    if tax_rate > 1.0:
        tax_rate /= 100.0

    # First/second discount periods stored across I/J/K columns
    def _ij_float(key: str, default: float = 0.0) -> float:
        key_lower = key.lower()
        for k, v in ij_map.items():
            if key_lower in k.lower():
                try:
                    return float(v)
                except (TypeError, ValueError):
                    pass
        return default

    tax_holiday_marker = int(_ij_float("tax holiday", default=0.0))
    first_discount_marker = int(_ij_float("first discount", default=9.0))
    first_discount_rate = _find_float("first discount rate", default=0.05)
    if first_discount_rate > 1.0:
        first_discount_rate /= 100.0
    second_discount_marker = int(_ij_float("second discount", default=0.0))
    second_discount_rate = _find_float("second discount rate", default=0.10)
    if second_discount_rate > 1.0:
        second_discount_rate /= 100.0

    # Workbook tax rows store period end markers in column J, not durations.
    # Example: Tax Holiday J=5 with rate K=0 means years 1-4 are exempt and the
    # first reduced-rate period starts in year 5.
    tax_holiday_years = max(tax_holiday_marker - 1, 0)
    first_discount_years = max(first_discount_marker - tax_holiday_marker, 0)
    second_discount_years = max(second_discount_marker - first_discount_marker, 0)

    # --- MRA parameters (Assumption!K46–K47, K35–K36, Other Input rows 3–6) ---
    bess_mra_pct = _find_float("bess mra", default=0.60)
    if bess_mra_pct > 1.0:
        bess_mra_pct /= 100.0
    pv_mra_pct = _find_float("pv mra", default=0.10)
    if pv_mra_pct > 1.0:
        pv_mra_pct /= 100.0

    # MRA buildup schedule: load maintenance years + build-up fractions from workbook.
    # Assumption!L35 = PV maintenance year (0 = not scheduled).
    # Assumption!L36 = BESS maintenance year (0 = not scheduled).
    # Other Input rows 3–N, cols B (year offset from maintenance) and C (fraction).
    # Year offset 0 = the maintenance year itself, offset k = k years before maintenance.
    mra_buildup_schedule: dict[int, float] = {}
    pv_maintenance_year = 0
    bess_maintenance_year = 0
    try:
        wb_mra = load_workbook(str(path), data_only=True, read_only=True)
        ws_assum = wb_mra[ASSUMPTIONS_SHEET]

        # Read maintenance years from col K (col 11) in Assumption (label in col I)
        pv_maintenance_year = 0
        bess_maintenance_year = 0
        for row in range(1, (ws_assum.max_row or 1) + 1):
            label = str(ws_assum.cell(row=row, column=9).value or "").lower()  # col I
            val = ws_assum.cell(row=row, column=11).value  # col K
            if "pv maintenance" in label and val is not None:
                try:
                    pv_maintenance_year = int(float(val))
                except (TypeError, ValueError):
                    pass
            elif "bess maintenance" in label and val is not None:
                try:
                    bess_maintenance_year = int(float(val))
                except (TypeError, ValueError):
                    pass

        # Read build-up fractions from Other Input sheet
        if OTHER_INPUT_SHEET in wb_mra.sheetnames:
            ws_other = wb_mra[OTHER_INPUT_SHEET]
            buildup_fractions: dict[int, float] = {}  # offset → fraction
            for row in range(2, 20):
                offset_val = ws_other.cell(row=row, column=2).value  # col B = offset year
                pct_val = ws_other.cell(row=row, column=3).value  # col C = fraction
                if offset_val is None and pct_val is None:
                    continue
                try:
                    offset = int(float(offset_val))
                    pct = float(pct_val)
                    if 0 <= pct <= 1 and offset >= 0:
                        buildup_fractions[offset] = pct
                except (TypeError, ValueError):
                    continue

            # Convert (maintenance_year, offset) → absolute project year
            if pv_maintenance_year > 0 and buildup_fractions:
                for offset, pct in buildup_fractions.items():
                    abs_year = pv_maintenance_year - offset
                    if abs_year > 0:
                        mra_buildup_schedule[abs_year] = (
                            mra_buildup_schedule.get(abs_year, 0.0) + pct
                        )
            if bess_maintenance_year > 0 and buildup_fractions:
                for offset, pct in buildup_fractions.items():
                    abs_year = bess_maintenance_year - offset
                    if abs_year > 0:
                        mra_buildup_schedule[abs_year] = (
                            mra_buildup_schedule.get(abs_year, 0.0) + pct
                        )

        wb_mra.close()
    except Exception as exc:
        logger.warning("MRA buildup schedule load failed: %s — using default 4-year schedule", exc)
        mra_buildup_schedule = {}  # empty → build_mra_schedule uses its own default

    # --- PPA / revenue scenario parameters (Assumption!O/Q) ---
    def _oq_float(key: str, default: float = 0.0) -> float:
        key_lower = key.lower()
        for k, v in oq_map.items():
            if key_lower in k.lower():
                try:
                    return float(v)
                except (TypeError, ValueError):
                    pass
        return default

    ppa_option = int(_oq_float("ppa setting", default=3.0))
    if ppa_option not in (1, 2, 3, 4):
        ppa_option = 3
    bundled_discount_pct = _oq_float("bundled discount", default=0.15)
    if bundled_discount_pct > 1.0:
        bundled_discount_pct /= 100.0
    pv_discount_pct = _oq_float("pv discount", default=0.05)
    if pv_discount_pct > 1.0:
        pv_discount_pct /= 100.0
    bess_discount_pct = _oq_float("bess discount", default=0.05)
    if bess_discount_pct > 1.0:
        bess_discount_pct /= 100.0
    fixed_ppa_price_usd_per_mwh = _oq_float("fixed ppa price", default=70.0)
    revenue_escalation_pct = _oq_float("price escalation", default=0.05)
    if revenue_escalation_pct > 1.0:
        revenue_escalation_pct /= 100.0
    fmp_descent_pct = _oq_float("market price descent", default=-0.05)
    if fmp_descent_pct < -1.0:
        fmp_descent_pct /= 100.0
    # Capacity/demand tariff (USD/MW/year, Financial!H32 = Assumption!Q57)
    capacity_demand_rate_usd_per_mw = _oq_float("capacity", default=0.0)

    return {
        "project_years": project_years,
        "interest_rate_pct": (base_rate + debt_margin) * 100.0,
        "tenor_years": tenor_years,
        "target_dscr": target_dscr,
        "initial_capex_usd": initial_capex_usd,
        "discount_rate_pct": min_equity_irr_ratio * 100.0,
        "cod_date": _find_date_iso("commercial operation date", default_iso="2027-01-01"),
        "max_leverage_ratio": max_leverage_ratio,
        "exchange_rate_usd_vnd": exchange_rate_usd_vnd,
        # CAPEX breakdown
        "solar_capex_usd": solar_capex,
        "bess_capex_usd": bess_capex,
        "bop_capex_usd": bop_capex,
        "land_capex_usd": land_capex,
        "installed_pv_mwp": installed_pv_mwp,
        "bess_mwh": bess_mwh,
        # OPEX parameters
        "om_solar_usd_per_mwp": om_solar_usd_per_mwp,
        "om_bess_usd_per_mwh": om_bess_usd_per_mwh,
        "insurance_solar_pct_capex": insurance_solar_pct,
        "insurance_bess_pct_capex": insurance_bess_pct,
        "other_opex_usd_per_mwp": other_opex_usd_per_mwp,
        "asset_management_usd_per_mwp": asset_mgmt_usd_per_mwp,
        "land_lease_pct_revenue": land_lease_pct_revenue,
        "opex_escalation_pct": opex_escalation_pct,
        # Tax parameters
        "depreciation_tenor_years": depreciation_tenor_years,
        "tax_rate": tax_rate,
        "tax_holiday_years": tax_holiday_years,
        "first_discount_years": first_discount_years,
        "first_discount_rate": first_discount_rate,
        "second_discount_years": second_discount_years,
        "second_discount_rate": second_discount_rate,
        # MRA parameters
        # Zero out BESS MRA when no BESS maintenance is scheduled (maintenance_year=0).
        "bess_mra_pct": bess_mra_pct if bess_maintenance_year > 0 else 0.0,
        "pv_mra_pct": pv_mra_pct,
        "mra_buildup_schedule": mra_buildup_schedule if mra_buildup_schedule else None,
        # PPA / scenario parameters
        "ppa_option": ppa_option,
        "bundled_discount_pct": bundled_discount_pct,
        "pv_discount_pct": pv_discount_pct,
        "bess_discount_pct": bess_discount_pct,
        "fixed_ppa_price_usd_per_mwh": fixed_ppa_price_usd_per_mwh,
        "revenue_escalation_pct": revenue_escalation_pct,
        "fmp_descent_pct": fmp_descent_pct,
        "capacity_demand_rate_usd_per_mw": capacity_demand_rate_usd_per_mw,
    }


def load_hourly_data(path: Path) -> HourlyTimeSeries:
    """
    Load and validate the hourly time series data.

    Args:
        path: Path to Excel input file.

    Returns:
        HourlyTimeSeries DataFrame.

    Raises:
        InputValidationError: If row count or required columns are invalid.
    """
    df = _read_data_input_sheet(path)

    # Try column rename for real Excel layout before validation
    df = _normalize_hourly_columns(df)

    if len(df) not in (HOURS_PER_YEAR, HOURS_PER_LEAP_YEAR):
        raise InputValidationError(
            f"Expected 8760 or 8784 rows, got {len(df)}. Check for leap year or incomplete data."
        )

    missing = _missing_columns(df, REQUIRED_HOURLY_COLUMNS)
    if missing:
        raise InputValidationError(f"Missing required hourly columns: {sorted(missing)}.")

    for column in ("simulation_profile_kw", "irradiation_wh_m2", "load_kw"):
        if (df[column] < 0).any():
            raise InputValidationError(f"Hourly column '{column}' contains negative values.")

    return df


def _normalize_header_cell(value: object) -> str:
    """Normalize header values for robust row detection."""
    if value is None:
        return ""
    return str(value).strip().lower()


def _find_header_row(
    ws: Any,
    markers: set[str],
    search_rows: int = 40,
    search_cols: int = 20,
) -> int | None:
    """Find the first row containing all marker labels."""
    for row in range(1, min((ws.max_row or 1), search_rows) + 1):
        cells = {
            _normalize_header_cell(ws.cell(row=row, column=col).value)
            for col in range(1, search_cols + 1)
        }
        if markers.issubset(cells):
            return row
    return None


def _read_data_input_sheet(path: Path) -> pd.DataFrame:
    """
    Read Data Input with dynamic header-row detection.

    Why: Newer workbooks include metadata rows above the actual hourly header,
    so fixed-header pandas reads produce extra non-hourly rows.
    """
    try:
        wb = load_workbook(str(path), data_only=True, read_only=False)
    except (FileNotFoundError, OSError) as exc:
        raise InputValidationError(f"Failed to open Excel file {path}: {exc}") from exc

    if DATA_INPUT_SHEET not in wb.sheetnames:
        wb.close()
        raise InputValidationError(f"Sheet '{DATA_INPUT_SHEET}' not found in {path}.")

    ws = wb[DATA_INPUT_SHEET]
    header_row = _find_header_row(ws, _DATA_INPUT_HEADER_MARKERS)
    wb.close()

    if header_row is None:
        # Fall back to legacy behavior for older fixtures.
        return _read_sheet(path, DATA_INPUT_SHEET)

    try:
        df = pd.read_excel(path, sheet_name=DATA_INPUT_SHEET, header=header_row - 1)
    except (FileNotFoundError, OSError, ValueError) as exc:
        raise InputValidationError(
            f"Failed to read sheet '{DATA_INPUT_SHEET}' from {path}: {exc}"
        ) from exc

    # Keep only rows that look like real hourly timesteps.
    if "DateTime" in df.columns:
        datetime_series = pd.to_datetime(df["DateTime"], errors="coerce")
        df = df.loc[datetime_series.notna()].copy()

    return df.reset_index(drop=True)


# Map from real Excel column names (case-insensitive) to internal names.
# The real Data Input sheet uses: DateTime, SimulationProfile_kW,
# Irradiation_W/m2, Load_kW, FMP, CFMP.
_HOURLY_COLUMN_ALIASES: dict[str, str] = {
    "datetime": "datetime",
    "simulationprofile_kw": "simulation_profile_kw",
    "irradiation_w/m2": "irradiation_wh_m2",
    "load_kw": "load_kw",
    "fmp": "fmp_usd_per_kwh",
    "cfmp": "cfmp_usd_per_kwh",
}


def _normalize_hourly_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rename real Excel column names to internal conventions.

    Why: The real Data Input sheet uses column names like
    'SimulationProfile_kW' and 'FMP', while the physics engine expects
    'simulation_profile_kw' and 'fmp_usd_per_kwh'. This bridges the gap
    without requiring the Excel files to be preprocessed.

    Args:
        df: Raw DataFrame from the Data Input sheet.

    Returns:
        DataFrame with normalized column names. Does not mutate the input.
    """
    # If columns already match internal names, return as-is
    if REQUIRED_HOURLY_COLUMNS.issubset(set(df.columns)):
        return df

    rename_map: dict[str, str] = {}
    for col in df.columns:
        col_lower = str(col).strip().lower()
        if col_lower in _HOURLY_COLUMN_ALIASES:
            target = _HOURLY_COLUMN_ALIASES[col_lower]
            if target not in df.columns:
                rename_map[col] = target

    if rename_map:
        logger.info("Renaming Data Input columns: %s", rename_map)
        df = df.rename(columns=rename_map)

    return df


def load_degradation_table(path: Path, project_years: int = 25) -> pd.DataFrame:
    """
    Load and validate the degradation (Loss) table.

    Args:
        path: Path to Excel input file.
        project_years: Expected project length in years.

    Returns:
        DataFrame with degradation factors.

    Raises:
        InputValidationError: If columns or values are invalid.
        DegradationTableError: If year coverage is incomplete.
    """
    df = _read_loss_sheet(path)

    missing = _missing_columns(df, REQUIRED_DEGRADATION_COLUMNS)
    if missing:
        raise InputValidationError(f"Missing required degradation columns: {sorted(missing)}.")

    invalid_mask = (
        (df["pv_factor"] <= 0)
        | (df["pv_factor"] > 1)
        | (df["battery_factor_no_replacement"] <= 0)
        | (df["battery_factor_no_replacement"] > 1)
        | (df["battery_factor_with_replacement"] <= 0)
        | (df["battery_factor_with_replacement"] > 1)
    )
    if invalid_mask.any():
        raise InputValidationError("Degradation factors out of range (0, 1].")

    years = set(df["year"].astype(int).tolist())
    missing_years = [year for year in range(1, project_years + 1) if year not in years]
    if missing_years:
        raise DegradationTableError(
            f"Missing degradation years: {missing_years}", missing_years=missing_years
        )

    return df


# Map from real Loss sheet column names (case-insensitive) to internal names.
# Real headers: "Year ", "Battery's Loss", "Battery", "PV Loss", "PV",
# "Battery wt Replacement"
_LOSS_COLUMN_ALIASES: dict[str, str] = {
    "year": "year",
    "pv": "pv_factor",
    "pv cumulative retention": "pv_factor",
    "battery": "battery_factor_no_replacement",
    "bess cumulative retention": "battery_factor_no_replacement",
    "battery wt replacement": "battery_factor_with_replacement",
    "bess w/ replacement": "battery_factor_with_replacement",
}


def _read_loss_sheet(path: Path) -> pd.DataFrame:
    """
    Read and normalize the Loss sheet from an Excel file.

    Why: The real Loss sheet has a header row that pandas may not
    auto-detect (row 2 in openpyxl = row 1 in the sheet header).
    Column names like 'PV' and 'Battery' need mapping to internal
    names like 'pv_factor' and 'battery_factor_no_replacement'.

    Args:
        path: Path to Excel input file.

    Returns:
        DataFrame with normalized column names.

    Raises:
        InputValidationError: If the sheet cannot be loaded.
    """
    df = _read_sheet(path, LOSS_SHEET)

    # If columns already match internal names, return as-is
    if REQUIRED_DEGRADATION_COLUMNS.issubset(set(df.columns)):
        return df

    # The real Loss sheet has "Loss Table " as column A header (row 1 in Excel).
    # The actual column headers are in the first data row. Check if the first
    # column looks like a header row.
    first_col = str(df.columns[0]).strip().lower()
    if "loss" in first_col or "unnamed" in first_col:
        try:
            wb = load_workbook(str(path), data_only=True, read_only=False)
        except (FileNotFoundError, OSError) as exc:
            raise InputValidationError(f"Failed to open Excel file {path}: {exc}") from exc

        if LOSS_SHEET not in wb.sheetnames:
            wb.close()
            raise InputValidationError(f"Sheet '{LOSS_SHEET}' not found in {path}.")

        ws = wb[LOSS_SHEET]
        header_row = _find_header_row(ws, _LOSS_HEADER_MARKERS)
        wb.close()

        if header_row is None:
            header_row = 2

        try:
            df = pd.read_excel(path, sheet_name=LOSS_SHEET, header=header_row - 1)
        except (FileNotFoundError, OSError, ValueError) as exc:
            raise InputValidationError(f"Failed to re-read Loss sheet from {path}: {exc}") from exc

    # Normalize column names
    rename_map: dict[str, str] = {}
    for col in df.columns:
        col_lower = str(col).strip().lower()
        if col_lower in _LOSS_COLUMN_ALIASES:
            target = _LOSS_COLUMN_ALIASES[col_lower]
            if target not in df.columns and target not in rename_map.values():
                rename_map[col] = target

    if rename_map:
        logger.info("Renaming Loss columns: %s", rename_map)
        df = df.rename(columns=rename_map)

    # Drop rows where year is NaN (possible trailing empty rows)
    if "year" in df.columns:
        year_numeric = pd.to_numeric(df["year"], errors="coerce")
        df = df.loc[year_numeric.notna()].copy()
        df["year"] = year_numeric.loc[df.index].astype(int)

    return df


def load_tariff_schedule_from_calc(path: Path) -> dict[TimePeriod, list[int]]:
    """
    Load tariff schedule from the Calc sheet's TimePeriodFlag column ('O', 'S', 'P').

    The Calc sheet assigns each hour a time-period flag based on the Vietnamese EVN
    tariff schedule. Reading the first 24 rows (one per hour of day) captures the
    full 24-hour cycle.
    """
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except (FileNotFoundError, OSError) as exc:
        raise InputValidationError(f"Failed to open {path}: {exc}") from exc

    if CALC_SHEET not in wb.sheetnames:
        wb.close()
        raise InputValidationError(f"Sheet '{CALC_SHEET}' not found in {path}.")

    ws = wb[CALC_SHEET]
    # Row 1 = header, row 2 = hour 0, row 3 = hour 1, ..., row 25 = hour 23
    # Col E = TimePeriodFlag: 'O' = off-peak, 'S' = standard, 'P' = peak
    flag_col = None
    header_row = ws[1]
    for cell in header_row:
        if cell.value and "timeperiodflag" in str(cell.value).lower().replace(" ", ""):
            flag_col = cell.column
            break
    wb.close()

    if flag_col is None:
        raise InputValidationError("'TimePeriodFlag' column not found in Calc sheet.")

    wb2 = load_workbook(str(path), data_only=True, read_only=True)
    ws2 = wb2[CALC_SHEET]

    # Calc sheet uses 'O'=off-peak, 'N'=normal/standard, 'P'=peak (and legacy 'S'=standard)
    flag_map = {
        "o": TimePeriod.OFF_PEAK,
        "s": TimePeriod.STANDARD,
        "n": TimePeriod.STANDARD,
        "p": TimePeriod.PEAK,
    }
    schedule: dict[TimePeriod, list[int]] = {
        TimePeriod.OFF_PEAK: [],
        TimePeriod.STANDARD: [],
        TimePeriod.PEAK: [],
    }
    for hour in range(24):
        row_idx = hour + 2  # row 2 = hour 0
        raw = ws2.cell(row=row_idx, column=flag_col).value
        if raw is None:
            continue
        flag = str(raw).strip().lower()[0] if str(raw).strip() else ""
        period = flag_map.get(flag)
        if period is not None:
            schedule[period].append(hour)

    wb2.close()
    return schedule


def load_tariff_schedule(path: Path) -> dict[TimePeriod, list[int]]:
    """
    Load tariff schedule defining peak/off-peak hours.

    Args:
        path: Path to Excel input file.

    Returns:
        Mapping from TimePeriod to list of hours.

    Raises:
        InputValidationError: If the schedule contains invalid hours or periods.
    """
    df = _read_sheet(path, TARIFF_SHEET)

    if _missing_columns(df, {"hour", "period"}):
        raise InputValidationError("Tariff schedule must contain 'hour' and 'period'.")

    if (df["hour"] < 0).any() or (df["hour"] > 23).any():
        raise InputValidationError("Invalid hour in tariff schedule (must be 0-23).")

    period_map = {
        "off_peak": TimePeriod.OFF_PEAK,
        "standard": TimePeriod.STANDARD,
        "peak": TimePeriod.PEAK,
    }

    schedule: dict[TimePeriod, list[int]] = {
        TimePeriod.OFF_PEAK: [],
        TimePeriod.STANDARD: [],
        TimePeriod.PEAK: [],
    }

    for _, row in df.iterrows():
        period_key = str(row["period"]).strip().lower()
        if period_key not in period_map:
            raise InputValidationError(
                f"Invalid tariff period '{row['period']}'. Expected off_peak, standard, peak."
            )
        period = period_map[period_key]
        schedule[period].append(int(row["hour"]))

    return schedule


def _read_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    """
    Read a sheet from an Excel file with standard error handling.

    Args:
        path: Path to Excel input file.
        sheet_name: Name of sheet to read.

    Returns:
        DataFrame for the sheet.

    Raises:
        InputValidationError: If sheet cannot be loaded.
    """
    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except (FileNotFoundError, OSError, ValueError) as exc:  # pragma: no cover - IO errors
        raise InputValidationError(
            f"Failed to read sheet '{sheet_name}' from {path}: {exc}"
        ) from exc


def _missing_columns(df: pd.DataFrame, required: set[str]) -> set[str]:
    """
    Return the set of missing required columns for a DataFrame.
    """
    return set(required) - set(df.columns)
