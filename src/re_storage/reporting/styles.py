"""Allotrope brand style constants for Excel workbook formatting.

Reusable fonts, fills, borders, and conditional-formatting colors used across
all reporting sheets.
"""

from __future__ import annotations

from openpyxl.styles import Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Brand colors
# ---------------------------------------------------------------------------

BRAND_GREEN = "2E7D32"
BRAND_GREEN_DARK = "1B5E20"
BRAND_GREEN_LIGHT = "E8F5E9"
BRAND_TEXT_DARK = "212121"
BRAND_TEXT_LIGHT = "757575"
BRAND_BLUE = "1565C0"
BRAND_YELLOW = "F9A825"
BRAND_GRAY = "9E9E9E"
BRAND_RED = "C62828"

# ---------------------------------------------------------------------------
# Conditional formatting fills
# ---------------------------------------------------------------------------

PASS_FILL = PatternFill("solid", fgColor="C8E6C9")
CAUTION_FILL = PatternFill("solid", fgColor="FFE0B2")
FAIL_FILL = PatternFill("solid", fgColor="FFCDD2")
BEST_FILL = PatternFill("solid", fgColor="C6EFCE")

# ---------------------------------------------------------------------------
# Fonts
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri Light", size=12, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri Light", size=16, bold=True, color=BRAND_GREEN_DARK)
BODY_FONT = Font(name="Calibri", size=10)
BODY_BOLD = Font(name="Calibri", size=10, bold=True)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=BRAND_GREEN_DARK)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color=BRAND_TEXT_LIGHT)
CONFIDENTIAL_FONT = Font(name="Calibri", size=8, italic=True, color=BRAND_TEXT_LIGHT)

# ---------------------------------------------------------------------------
# Fills
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor=BRAND_GREEN)
SECTION_FILL = PatternFill("solid", fgColor=BRAND_GREEN_LIGHT)
ALT_ROW_FILL = PatternFill("solid", fgColor="F5F5F5")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")

# ---------------------------------------------------------------------------
# Borders
# ---------------------------------------------------------------------------

THIN_BORDER = Border(bottom=Side(style="thin", color="BDBDBD"))
TOTAL_BORDER = Border(top=Side(style="medium", color=BRAND_GREEN))
SECTION_BORDER = Border(
    top=Side(style="medium", color=BRAND_GREEN),
    bottom=Side(style="thin", color="BDBDBD"),
)
HEADER_BORDER = Border(
    bottom=Side(style="medium", color=BRAND_GREEN),
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
)
