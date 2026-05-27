"""
Excel workbook writer for DPPA assessment deliverables.

Produces multi-sheet formatted .xlsx workbooks from pipeline outputs using openpyxl.
All values are static (no Excel formulas or VBA macros) — client reviews numbers,
not a live model.
"""

from __future__ import annotations

import logging
import math
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from re_storage.reporting.styles import (
    ALT_ROW_FILL,
    BEST_FILL,
    BODY_BOLD,
    BODY_FONT,
    CAUTION_FILL,
    CONFIDENTIAL_FONT,
    FAIL_FILL,
    HEADER_BORDER,
    HEADER_FILL,
    HEADER_FONT,
    NOTE_FONT,
    PASS_FILL,
    SECTION_BORDER,
    SECTION_FILL,
    SECTION_FONT,
    THIN_BORDER,
    TITLE_FONT,
    TOTAL_BORDER,
    TOTAL_FILL,
    TOTAL_FONT,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Number format mapping
# ---------------------------------------------------------------------------

_NUMBER_FORMATS: dict[str, str] = {
    "irr": "0.00%",
    "pct": "0.00%",
    "usd": '#,##0',
    "dscr": "0.00",
    "years": "0.0",
    "mwh": "#,##0.0",
    "kwh": "#,##0.0",
    "ratio": "0.00",
    "fx_rate": "#,##0.00",
}


def _apply_number_format(cell: openpyxl.cell.Cell, key: str) -> None:
    """Map a KPI key name to an Excel number format string."""
    key_lower = key.lower()
    for pattern, fmt in _NUMBER_FORMATS.items():
        if pattern in key_lower:
            cell.number_format = fmt
            return


def _format_value(value: Any, key: str) -> Any:
    """Sanitize a value for Excel display — convert NaN/Inf to 'N/A'."""
    if value is None:
        return "N/A"
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return "N/A"
    return value


def _auto_width(ws: openpyxl.worksheet.worksheet.Worksheet, max_width: int = 35) -> None:
    """Estimate column widths from content and apply them."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _write_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    headers: list[str],
) -> None:
    """Write a formatted header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER


def _write_data_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    values: list[Any],
    keys: list[str] | None = None,
    alt: bool = False,
) -> None:
    """Write a formatted data row with optional number formatting."""
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="center")
        if alt:
            cell.fill = ALT_ROW_FILL
        if keys and col_idx <= len(keys):
            _apply_number_format(cell, keys[col_idx - 1])


def _write_section_header(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    text: str,
    max_col: int = 4,
) -> None:
    """Write a section header spanning multiple columns."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.border = SECTION_BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def create_workbook() -> openpyxl.Workbook:
    """Create a new workbook with default styling and no default sheet."""
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)
    wb.properties.creator = "RE-Storage Model"
    wb.properties.created = datetime.now()
    return wb


def write_cover_sheet(
    wb: openpyxl.Workbook,
    project_name: str,
    project_metadata: dict[str, Any],
    kpis: dict[str, Any],
    verdict: Any | None = None,
) -> None:
    """Write the cover sheet with project name, date, KPI summary, and optional verdict."""
    from re_storage.reporting.assessment import AssessmentVerdict

    ws = wb.create_sheet("Cover")

    # Title
    ws.merge_cells("A1:B1")
    cell = ws.cell(row=1, column=1, value=project_name)
    cell.font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    # Date and confidentiality
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=2, column=1).font = NOTE_FONT
    ws.cell(row=3, column=1, value="CONFIDENTIAL — Prepared by Allotrope Ventures")
    ws.cell(row=3, column=1).font = CONFIDENTIAL_FONT

    # KPI summary table
    ws.cell(row=5, column=1, value="Key Performance Indicators")
    ws.cell(row=5, column=1).font = SECTION_FONT
    ws.cell(row=5, column=1).fill = SECTION_FILL

    _write_header_row(ws, 6, ["Metric", "Value"])

    display_order = [
        "project_irr",
        "equity_irr",
        "unlevered_irr",
        "npv_usd",
        "after_tax_npv_usd",
        "dscr_min",
        "dscr_avg",
        "debt_amount_usd",
        "simple_payback_years",
        "discounted_payback_year",
        "cash_on_cash_yield",
        "year1_solar_generation_mwh",
        "year1_dppa_revenue_usd",
        "year1_grid_savings_usd",
        "year1_opex_usd",
        "year1_ebitda_usd",
    ]

    label_map = {
        "project_irr": "Project IRR",
        "equity_irr": "Equity IRR",
        "unlevered_irr": "Unlevered IRR",
        "npv_usd": "NPV (USD)",
        "after_tax_npv_usd": "After-Tax NPV (USD)",
        "dscr_min": "Min DSCR",
        "dscr_avg": "Avg DSCR",
        "debt_amount_usd": "Debt Amount (USD)",
        "simple_payback_years": "Simple Payback (Years)",
        "discounted_payback_year": "Discounted Payback (Year Index)",
        "cash_on_cash_yield": "Cash-on-Cash Yield",
        "year1_solar_generation_mwh": "Year 1 Solar Generation (MWh)",
        "year1_dppa_revenue_usd": "Year 1 DPPA Revenue (USD)",
        "year1_grid_savings_usd": "Year 1 Grid Savings (USD)",
        "year1_opex_usd": "Year 1 OPEX (USD)",
        "year1_ebitda_usd": "Year 1 EBITDA (USD)",
    }

    row = 7
    for key in display_order:
        if key not in kpis:
            continue
        label = label_map.get(key, key)
        value = _format_value(kpis[key], key)
        _write_data_row(ws, row, [label, value], keys=["", key], alt=(row % 2 == 0))
        row += 1

    # Assessment verdict section
    if verdict is not None and isinstance(verdict, AssessmentVerdict):
        row += 1
        ws.cell(row=row, column=1, value="ASSESSMENT VERDICT")
        ws.cell(row=row, column=1).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

        overall_fill = {
            "GO": PASS_FILL,
            "CAUTION": CAUTION_FILL,
            "NO-GO": FAIL_FILL,
        }.get(verdict.overall, PatternFill())

        cell = ws.cell(row=row, column=1, value=f"Overall: {verdict.overall}")
        cell.font = BODY_BOLD
        cell.fill = overall_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

        status_map = {
            "Equity IRR": verdict.equity_irr_status,
            "Min DSCR": verdict.dscr_status,
            "NPV": verdict.npv_status,
            "Payback": verdict.payback_status,
        }
        for label, status in status_map.items():
            status_fill = {
                "PASS": PASS_FILL,
                "MARGINAL": CAUTION_FILL,
                "FAIL": FAIL_FILL,
            }.get(status, PatternFill())
            cell = ws.cell(row=row, column=1, value=f"{label}: {status}")
            cell.font = BODY_FONT
            cell.fill = status_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1

        for detail in verdict.details:
            cell = ws.cell(row=row, column=1, value=detail)
            cell.font = NOTE_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1

    # Metadata section
    row += 1
    ws.cell(row=row, column=1, value="Project Metadata")
    ws.cell(row=row, column=1).font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1
    for meta_key, meta_value in project_metadata.items():
        _write_data_row(ws, row, [meta_key, meta_value], alt=(row % 2 == 0))
        row += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20


def write_assessment_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    kpis: dict[str, Any],
    annual_df: Any,
    exchange_rate_usd_vnd: float = 25_000.0,
    charts: list[Any] | None = None,
) -> None:
    """Write an assessment sheet with KPI cards, annual proforma table, and optional charts."""
    import pandas as pd

    ws = wb.create_sheet(sheet_name)

    # Section 1: KPI cards
    ws.cell(row=1, column=1, value="Key Performance Indicators")
    ws.cell(row=1, column=1).font = SECTION_FONT
    ws.cell(row=1, column=1).fill = SECTION_FILL

    _write_header_row(ws, 2, ["Metric", "USD", "VND"])

    usd_keys = [k for k in kpis if k.endswith("_usd") or k.endswith("_irr") or k.endswith("_yield")]
    row = 3
    for key in usd_keys[:12]:
        label = key.replace("_", " ").title()
        usd_val = _format_value(kpis[key], key)
        vnd_val = "N/A"
        if isinstance(usd_val, (int, float)) and exchange_rate_usd_vnd > 0:
            vnd_val = round(usd_val * exchange_rate_usd_vnd, 0)
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        cell_usd = ws.cell(row=row, column=2, value=usd_val)
        cell_usd.font = BODY_FONT
        _apply_number_format(cell_usd, key)
        cell_vnd = ws.cell(row=row, column=3, value=vnd_val)
        cell_vnd.font = BODY_FONT
        if isinstance(vnd_val, (int, float)):
            cell_vnd.number_format = '#,##0'
        if row % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = ALT_ROW_FILL
        row += 1

    # Section 2: Annual proforma
    if annual_df is None or (isinstance(annual_df, pd.DataFrame) and annual_df.empty):
        pass
    else:
        row += 1
        _write_section_header(ws, row, "Annual Proforma", max_col=5)
        row += 1

        # Headers with VND columns
        usd_cols = [c for c in annual_df.columns if c.endswith("_usd") or c == "dscr" or c == "year"]
        headers = ["Year"]
        keys = ["year"]
        for col in usd_cols:
            if col == "year":
                continue
            headers.append(col.replace("_", " ").title())
            headers.append(col.replace("_", " ").title() + " (VND)")
            keys.append(col)
            keys.append(col)

        _write_header_row(ws, row, headers)
        row += 1

        for _, record in annual_df.iterrows():
            values = [int(record.get("year", 0))]
            for col in usd_cols:
                if col == "year":
                    continue
                raw_val = _format_value(record.get(col), col)
                values.append(raw_val)
                if isinstance(raw_val, (int, float)) and exchange_rate_usd_vnd > 0:
                    values.append(round(raw_val * exchange_rate_usd_vnd, 0))
                else:
                    values.append("N/A")
            _write_data_row(ws, row, values, keys=keys, alt=(row % 2 == 0))
            row += 1

    # Section 3: Embedded charts
    if charts:
        row += 2
        try:
            from openpyxl.drawing.image import Image as XLImage

            for chart_path in charts:
                if chart_path is None:
                    continue
                img = XLImage(str(chart_path))
                # Scale to fit roughly columns A-L (~800 px wide)
                img.width = 800
                img.height = 400
                ws.add_image(img, f"A{row}")
                row += 26  # Approximate rows for 400px height
        except Exception as exc:
            logger.warning("Could not embed charts: %s", exc)

    _auto_width(ws)


def write_comparison_sheet(
    wb: openpyxl.Workbook,
    scenario_results: dict[int, dict[str, Any]],
    exchange_rate_usd_vnd: float = 25_000.0,
) -> None:
    """Write a comparison sheet with scenario results side-by-side."""
    ws = wb.create_sheet("Comparison")

    option_labels = {
        1: "Option 1: Bundled",
        2: "Option 2: Separate",
        3: "Option 3: DPPA CfD",
        4: "Option 4: Fixed EVN",
    }

    kpi_keys = [
        "project_irr",
        "equity_irr",
        "npv_usd",
        "dscr_min",
        "year1_dppa_revenue_usd",
        "year1_grid_savings_usd",
        "year1_opex_usd",
        "year1_ebitda_usd",
        "simple_payback_years",
        "debt_amount_usd",
        "cash_on_cash_yield",
    ]

    label_map = {
        "project_irr": "Project IRR",
        "equity_irr": "Equity IRR",
        "npv_usd": "NPV (USD)",
        "dscr_min": "Min DSCR",
        "year1_dppa_revenue_usd": "Year 1 DPPA Revenue (USD)",
        "year1_grid_savings_usd": "Year 1 Grid Savings (USD)",
        "year1_opex_usd": "Year 1 OPEX (USD)",
        "year1_ebitda_usd": "Year 1 EBITDA (USD)",
        "simple_payback_years": "Simple Payback (Years)",
        "debt_amount_usd": "Debt Amount (USD)",
        "cash_on_cash_yield": "Cash-on-Cash Yield",
    }

    # Build column headers: Metric | Option1 USD | Option1 VND | Option2 USD | Option2 VND | ...
    headers = ["Metric"]
    for option_id in sorted(scenario_results.keys()):
        label = option_labels.get(option_id, f"Option {option_id}")
        headers.append(f"{label} (USD)")
        headers.append(f"{label} (VND)")

    _write_header_row(ws, 1, headers)

    # Find best value per row for highlighting
    for row_idx, kpi_key in enumerate(kpi_keys, 2):
        label = label_map.get(kpi_key, kpi_key)
        ws.cell(row=row_idx, column=1, value=label).font = BODY_BOLD

        # Collect numeric values for best-value highlighting
        numeric_values: list[tuple[int, float]] = []
        col_offset = 2
        for option_id in sorted(scenario_results.keys()):
            result = scenario_results[option_id]
            raw_val = _format_value(result.get(kpi_key), kpi_key)
            usd_cell = ws.cell(row=row_idx, column=col_offset, value=raw_val)
            usd_cell.font = BODY_FONT
            _apply_number_format(usd_cell, kpi_key)

            if isinstance(raw_val, (int, float)) and exchange_rate_usd_vnd > 0:
                vnd_val = round(raw_val * exchange_rate_usd_vnd, 0)
                numeric_values.append((col_offset, raw_val))
            else:
                vnd_val = "N/A"

            vnd_cell = ws.cell(row=row_idx, column=col_offset + 1, value=vnd_val)
            vnd_cell.font = BODY_FONT
            if isinstance(vnd_val, (int, float)):
                vnd_cell.number_format = '#,##0'
            col_offset += 2

        # Highlight best value (higher is better for IRR/NPV/DSCR, lower for payback)
        if numeric_values:
            higher_is_better = kpi_key not in (
                "simple_payback_years",
                "discounted_payback_year",
            )
            if higher_is_better:
                best_col = max(numeric_values, key=lambda x: x[1])[0]
            else:
                best_col = min(numeric_values, key=lambda x: x[1])[0]
            ws.cell(row=row_idx, column=best_col).fill = BEST_FILL

    _auto_width(ws)


def write_sensitivity_sheet(
    wb: openpyxl.Workbook,
    sensitivity_results: dict[str, dict[float, dict[str, Any]]],
    exchange_rate_usd_vnd: float = 25_000.0,
) -> None:
    """Write a sensitivity sheet with results per variable (top 5 by default)."""
    ws = wb.create_sheet("Sensitivity")

    # Only show top 5 variables to keep the sheet readable
    top_variables = [
        "strike_price",
        "interest_rate",
        "initial_capex",
        "exchange_rate",
        "bundled_discount",
    ]

    kpi_keys = ["project_irr", "equity_irr", "npv_usd", "dscr_min", "simple_payback_years"]

    label_map = {
        "project_irr": "Project IRR",
        "equity_irr": "Equity IRR",
        "npv_usd": "NPV (USD)",
        "dscr_min": "Min DSCR",
        "simple_payback_years": "Simple Payback (Years)",
    }

    row = 1
    for variable_name in top_variables:
        if variable_name not in sensitivity_results:
            continue

        results = sensitivity_results[variable_name]
        _write_section_header(ws, row, f"Sensitivity: {variable_name.replace('_', ' ').title()}", max_col=5)
        row += 1

        headers = ["Test Value"]
        for kpi_key in kpi_keys:
            headers.append(label_map.get(kpi_key, kpi_key))
            headers.append(label_map.get(kpi_key, kpi_key) + " (VND)")

        _write_header_row(ws, row, headers)
        row += 1

        for test_value, kpis in sorted(results.items()):
            values = [test_value]
            for kpi_key in kpi_keys:
                raw_val = _format_value(kpis.get(kpi_key), kpi_key)
                values.append(raw_val)
                if isinstance(raw_val, (int, float)) and exchange_rate_usd_vnd > 0:
                    values.append(round(raw_val * exchange_rate_usd_vnd, 0))
                else:
                    values.append("N/A")
            _write_data_row(ws, row, values, alt=(row % 2 == 0))
            row += 1

        row += 1

    _auto_width(ws)


def write_assumptions_sheet(
    wb: openpyxl.Workbook,
    assumptions_dict: dict[str, Any],
) -> None:
    """Write an assumptions sheet with parameter key-value pairs grouped by category."""
    ws = wb.create_sheet("Assumptions")

    _write_header_row(ws, 1, ["Parameter", "Value"])

    # Group by category
    categories: dict[str, list[tuple[str, Any]]] = {
        "System": [],
        "DPPA": [],
        "Financial": [],
        "Dispatch": [],
        "Other": [],
    }

    category_keywords = {
        "System": {"capacity", "bess", "solar", "pv", "degradation", "battery"},
        "DPPA": {"ppa", "strike", "dppa", "k_factor", "kpp", "bundled", "discount"},
        "Financial": {"capex", "interest", "tenor", "dscr", "tax", "debt", "leverage", "npv", "irr", "payback", "cash_on_cash", "exchange", "cpi", "escalation", "revenue"},
        "Dispatch": {"strategy", "charging", "peak", "when_needed", "after_sunset", "cycle", "optimize"},
    }

    for key, value in sorted(assumptions_dict.items()):
        placed = False
        for category, keywords in category_keywords.items():
            if any(kw in key.lower() for kw in keywords):
                categories[category].append((key, value))
                placed = True
                break
        if not placed:
            categories["Other"].append((key, value))

    row = 2
    for category_name, items in categories.items():
        if not items:
            continue
        _write_section_header(ws, row, category_name, max_col=2)
        row += 1
        for key, value in items:
            _write_data_row(ws, row, [key, _format_value(value, key)], alt=(row % 2 == 0))
            row += 1
        row += 1

    _auto_width(ws)


def save_workbook(wb: openpyxl.Workbook, output_path: str | Path) -> Path:
    """Save the workbook to disk."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Workbook saved to %s", output_path)
    return output_path
